from flask import Flask, render_template, request, jsonify, send_file, redirect
import sqlite3, os
from datetime import datetime, timedelta
import requests 
from zoneinfo import ZoneInfo
import csv, io
from calculos import calcular_comercial
from bs4 import BeautifulSoup
import re
from flask_login import (
    LoginManager,
    UserMixin,
    login_user,
    login_required,
    logout_user,
    current_user
)
from werkzeug.security import generate_password_hash, check_password_hash

app = Flask(__name__)
app.secret_key = "super_clave_cambiar_en_produccion"

login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = "login"

# ======================
# DB PATH
# ======================
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DB_NAME = os.path.join(BASE_DIR, "silobolsas.db")

# ======================
# UTILIDADES
# ======================
def ahora():
    return datetime.now(ZoneInfo("America/Argentina/Buenos_Aires"))


def get_db():
    conn = sqlite3.connect(DB_NAME)
    conn.row_factory = sqlite3.Row
    return conn

# ======================
# DB INIT
# ======================
def init_db():
    conn = get_db()
    c = conn.cursor()

    # TABLA USUARIOS
    c.execute("""
        CREATE TABLE IF NOT EXISTS usuarios (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            username TEXT UNIQUE,
            password TEXT,
            rol TEXT
        )
    """)
    # TABLA PERMISOS
    c.execute("""
        CREATE TABLE IF NOT EXISTS permisos (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER,
            pantalla TEXT
        )
    """)

    admin = c.execute(
        "SELECT * FROM usuarios WHERE username='admin'"
    ).fetchone()

    if not admin:
        c.execute("""
            INSERT INTO usuarios (username, password, rol)
            VALUES (?, ?, ?)
        """, (
            "admin",
            generate_password_hash("admin123"),
            "admin"
        ))

    # TABLAS ORIGINALES
    c.execute("""
        CREATE TABLE IF NOT EXISTS silos (
            numero_qr TEXT PRIMARY KEY,
            cereal TEXT,
            estado_grano TEXT,
            estado_silo TEXT,
            metros INTEGER,
            lat REAL,
            lon REAL,
            fecha_confeccion TEXT,
            fecha_extraccion TEXT
        )
    """)

    c.execute("""
        CREATE TABLE IF NOT EXISTS muestreos (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            numero_qr TEXT,
            fecha_muestreo TEXT
        )
    """)

    c.execute("""
        CREATE TABLE IF NOT EXISTS analisis (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            id_muestreo INTEGER,
            seccion TEXT,
            temperatura REAL,
            humedad REAL,
            ph REAL,
            danados REAL,
            quebrados REAL,
            materia_extrana REAL,
            olor REAL,
            moho REAL,
            insectos INTEGER,
            chamico REAL,
            grado INTEGER,
            factor REAL,
            tas INTEGER
        )
    """)

    c.execute("""
        CREATE TABLE IF NOT EXISTS monitoreos (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            numero_qr TEXT,
            fecha_evento TEXT,
            tipo TEXT,
            detalle TEXT,
            foto_evento TEXT,
            resuelto INTEGER DEFAULT 0,
            fecha_resolucion TEXT,
            foto_resolucion TEXT
        )
    """)
    c.execute("""
        CREATE TABLE IF NOT EXISTS solicitudes (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER,
            pantalla TEXT,
            fecha TEXT,
            estado TEXT DEFAULT 'pendiente'
        )
    """)

    c.execute("""
        CREATE TABLE IF NOT EXISTS mercado (
            cereal TEXT PRIMARY KEY,
            pizarra_auto REAL,
            pizarra_manual REAL,
            usar_manual INTEGER DEFAULT 0,
            obs_precio TEXT,
            dolar REAL,
            fecha TEXT,
            fuente TEXT,
            fecha_fuente TEXT
        )
    """)

# Asegurar columnas nuevas en Render
    try:
        c.execute("ALTER TABLE mercado ADD COLUMN fuente TEXT")
    except:
        pass

    try:
        c.execute("ALTER TABLE mercado ADD COLUMN fecha_fuente TEXT")
    except:
        pass

    # Datos iniciales mercado
    cereales_base = ["Soja", "Maíz", "Trigo", "Girasol"]

    for cereal in cereales_base:
        c.execute("""
            INSERT OR IGNORE INTO mercado (
                cereal,
                pizarra_auto,
                pizarra_manual,
                usar_manual,
                dolar,
                fecha
            )
        VALUES (?,?,?,?,?,?)
    """, (
        cereal,
        0,      # pizarra_auto
        0,      # pizarra_manual
        0,      # usar_manual
        0,      # dolar
        ahora().strftime("%Y-%m-%d %H:%M")
    ))

    conn.commit()
    conn.close()

init_db()

# ======================
# MODELO USUARIO
# ======================

class User(UserMixin):
    def __init__(self, id, username, rol):
        self.id = id
        self.username = username
        self.rol = rol


@login_manager.user_loader
def load_user(user_id):
    conn = get_db()
    try:
        u = conn.execute(
            "SELECT * FROM usuarios WHERE id=?",
            (user_id,)
        ).fetchone()
    except:
        conn.close()
        return None

    conn.close()

    if u:
        return User(u["id"], u["username"], u["rol"])
    return None


# ======================
# PERMISOS
# ======================

def tiene_permiso(pantalla):

    if not current_user.is_authenticated:
        return False

    # ADMIN siempre tiene todo
    if current_user.username == "admin":
        return True

    conn = get_db()
    row = conn.execute("""
        SELECT 1 FROM permisos
        WHERE user_id=? AND pantalla=?
    """, (current_user.id, pantalla)).fetchone()
    conn.close()

    return row is not None


def acceso_denegado(pantalla):

    conn = get_db()

    ya_solicitado = conn.execute("""
        SELECT 1 FROM solicitudes
        WHERE user_id=? AND pantalla=? AND estado='pendiente'
    """, (current_user.id, pantalla)).fetchone()

    solicitud_enviada = False

    if request.method == "POST" and not ya_solicitado:
        conn.execute("""
            INSERT INTO solicitudes (user_id, pantalla, fecha)
            VALUES (?,?,?)
        """, (
            current_user.id,
            pantalla,
            ahora().strftime("%Y-%m-%d %H:%M")
        ))
        conn.commit()
        solicitud_enviada = True
    elif ya_solicitado:
        solicitud_enviada = True

    conn.close()

    return render_template(
        "no_autorizado.html",
        pantalla=pantalla,
        solicitud_enviada=solicitud_enviada
    ), 403

# ======================
# PANTALLA PERMISOS
# ======================
@app.route("/admin/usuarios")
@login_required
def admin_usuarios():
    if not tiene_permiso("admin"):
        return acceso_denegado("admin_usuarios")

    conn = get_db()
    solicitudes = conn.execute("""
    SELECT s.*, u.username
    FROM solicitudes s
    JOIN usuarios u ON u.id = s.user_id
    WHERE s.estado='pendiente'
""").fetchall()

    usuarios = conn.execute("SELECT * FROM usuarios").fetchall()
    permisos = conn.execute("SELECT * FROM permisos").fetchall()
    conn.close()

    permisos_set = set((p["user_id"], p["pantalla"]) for p in permisos)
    return render_template(
    "admin_usuarios.html",
    usuarios=usuarios,
    permisos=permisos,
    permisos_set=permisos_set,
    solicitudes=solicitudes
)
@app.route("/solicitar_acceso/<pantalla>", methods=["POST"])
@login_required
def solicitar_acceso(pantalla):

    conn = get_db()

    ya_solicitado = conn.execute("""
        SELECT 1 FROM solicitudes
        WHERE user_id=? AND pantalla=? AND estado='pendiente'
    """, (current_user.id, pantalla)).fetchone()

    if not ya_solicitado:
        conn.execute("""
            INSERT INTO solicitudes (user_id, pantalla, fecha)
            VALUES (?,?,?)
        """, (
            current_user.id,
            pantalla,
            ahora().strftime("%Y-%m-%d %H:%M")
        ))
        conn.commit()

    conn.close()

    return acceso_denegado(pantalla)

@app.route("/admin/crear_usuario", methods=["POST"])
@login_required
def crear_usuario():

    if not tiene_permiso("admin"):
        return acceso_denegado("admin_usuarios")

    username = request.form.get("username")
    password = request.form.get("password")
    rol = request.form.get("rol")

    if not username or not password:
        return redirect("/admin/usuarios")

    conn = get_db()

    try:
        conn.execute("""
            INSERT INTO usuarios (username, password, rol)
            VALUES (?, ?, ?)
        """, (
            username,
            generate_password_hash(password),
            rol
        ))
        conn.commit()
    except:
        conn.close()
        return "El usuario ya existe"

    conn.close()
    return redirect("/admin/usuarios")

@app.route("/admin/permisos", methods=["POST"])
@login_required
def guardar_permisos():
    if not tiene_permiso("admin"):
        return acceso_denegado("admin_usuarios")

    user_id = request.form.get("user_id")
    permisos = request.form.getlist("permisos")

    conn = get_db()
    conn.execute("DELETE FROM permisos WHERE user_id=?", (user_id,))

    for p in permisos:
        conn.execute(
            "INSERT INTO permisos (user_id, pantalla) VALUES (?,?)",
            (user_id, p)
        )

    conn.commit()
    conn.close()

    return redirect("/admin/usuarios")
@app.route("/admin/eliminar_usuario", methods=["POST"])
@login_required
def eliminar_usuario():

    if not tiene_permiso("admin"):
        return "No autorizado", 403

    user_id = request.form.get("user_id")

    conn = get_db()

    # No permitir borrar admin principal
    u = conn.execute(
        "SELECT username FROM usuarios WHERE id=?",
        (user_id,)
    ).fetchone()

    if u and u["username"] == "admin":
        conn.close()
        return "No se puede eliminar el administrador principal"

    conn.execute("DELETE FROM permisos WHERE user_id=?", (user_id,))
    conn.execute("DELETE FROM usuarios WHERE id=?", (user_id,))
    conn.commit()
    conn.close()

    return redirect("/admin/usuarios")
@app.route("/admin/aprobar_solicitud", methods=["POST"])
@login_required
def aprobar_solicitud():
    if not tiene_permiso("admin"):
        return "No autorizado", 403

    solicitud_id = request.form.get("id")

    conn = get_db()
    s = conn.execute(
        "SELECT * FROM solicitudes WHERE id=?",
        (solicitud_id,)
    ).fetchone()

    if s:
        # dar permiso
        conn.execute(
            "INSERT INTO permisos (user_id, pantalla) VALUES (?,?)",
            (s["user_id"], s["pantalla"])
        )

        conn.execute(
            "UPDATE solicitudes SET estado='aprobado' WHERE id=?",
            (solicitud_id,)
        )

    conn.commit()
    conn.close()

    return redirect("/admin/usuarios")
@app.route("/admin/rechazar_solicitud", methods=["POST"])
@login_required
def rechazar_solicitud():
    if not tiene_permiso("admin"):
        return "No autorizado", 403

    solicitud_id = request.form.get("id")

    conn = get_db()
    conn.execute(
        "UPDATE solicitudes SET estado='rechazado' WHERE id=?",
        (solicitud_id,)
    )
    conn.commit()
    conn.close()

    return redirect("/admin/usuarios")

# ======================
# LOGIN
# ======================
@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        username = request.form.get("username")
        password = request.form.get("password")

        conn = get_db()
        u = conn.execute(
            "SELECT * FROM usuarios WHERE username=?",
            (username,)
        ).fetchone()
        conn.close()

        if u and check_password_hash(u["password"], password):
            user = User(u["id"], u["username"], u["rol"])
            login_user(user)

            # 🔥 Redirección inteligente
            if tiene_permiso("panel"):
                return redirect("/panel")
            elif tiene_permiso("form"):
                return redirect("/form")
            elif tiene_permiso("comercial"):
                return redirect("/comercial")
            elif tiene_permiso("admin"):
                return redirect("/admin/usuarios")
            else:
                return "No tiene permisos asignados"

        return render_template("login.html", error="Credenciales incorrectas")

    return render_template("login.html")


@app.route("/logout")
@login_required
def logout():
    logout_user()
    return redirect("/login")

def actualizar_tabla_mercado():
    conn = get_db()

    columnas = [
        "fuente TEXT",
        "fecha_fuente TEXT",
        "usar_manual INTEGER DEFAULT 0",
        "obs_precio TEXT"
    ]

    for col in columnas:
        try:
            conn.execute(f"ALTER TABLE mercado ADD COLUMN {col}")
        except:
            pass

    conn.commit()
    conn.close()



# ======================
# API — CONSULTA SILO
# ======================
@app.route("/api/silo/<qr>")
@login_required
def api_silo(qr):
    if not tiene_permiso("form"):
        return acceso_denegado("panel")
    conn = get_db()

    s = conn.execute("""
        SELECT
            s.cereal,
            s.fecha_confeccion,
            s.estado_silo,
            (
              SELECT MAX(fecha_muestreo)
              FROM muestreos
              WHERE numero_qr = s.numero_qr
            ) AS ultimo_calado
        FROM silos s
        WHERE s.numero_qr=?
    """, (qr,)).fetchone()

    conn.close()

    if not s:
        return jsonify(existe=False)

    return jsonify(
        existe=True,
        cereal=s["cereal"],
        fecha_confeccion=s["fecha_confeccion"],
        estado_silo=s["estado_silo"],
        ultimo_calado=s["ultimo_calado"]
    )

# ======================
# PANEL (ULTRA ROBUSTO)
# ======================
@app.route("/")
@app.route("/panel")
@login_required
def panel():
    if not tiene_permiso("panel"):
        return acceso_denegado("panel")
    conn = get_db()

    silos = conn.execute("""
    SELECT s.*,
    (
        SELECT m.id
        FROM muestreos m
        WHERE m.numero_qr = s.numero_qr
        ORDER BY m.id DESC
        LIMIT 1
    ) ultimo_muestreo
    FROM silos s
    ORDER BY datetime(fecha_confeccion) DESC
""").fetchall()

    registros = []
    
    for s in silos:
        # Conteo de eventos pendientes
        eventos_pendientes = conn.execute("""
            SELECT COUNT(*) AS cant
            FROM monitoreos
            WHERE numero_qr = ?
              AND resuelto = 0
        """, (s["numero_qr"],)).fetchone()["cant"]
        
        grado = None
        factor = None
        tas_min = None
        fecha_extraccion_estimada = None

        try:
            if s["ultimo_muestreo"]:
                analisis = conn.execute("""
                    SELECT grado, factor, tas
                    FROM analisis
                    WHERE id_muestreo=?
                """, (s["ultimo_muestreo"],)).fetchall()

                grados = []
                factores = []
                tass = []

                for a in analisis:
                    # grado: solo si es numérico
                    try:
                        g = int(a["grado"])
                        grados.append(g)
                    except:
                        pass

                    # factor: solo positivos
                    try:
                        f = float(a["factor"])
                        if f > 0:
                            factores.append(f)
                    except:
                        pass

                    # tas: solo positivos
                    try:
                        t = int(a["tas"])
                        if t > 0:
                            tass.append(t)
                    except:
                        pass

                grado = max(grados) if grados else None
                factor = round(sum(factores) / len(factores), 4) if factores else None
                tas_min = min(tass) if tass else None

                if tas_min:
                    row = conn.execute(
                        "SELECT fecha_muestreo FROM muestreos WHERE id=?",
                        (s["ultimo_muestreo"],)
                    ).fetchone()

                    fm = None
                    if row and row["fecha_muestreo"]:
                        for fmt in ("%Y-%m-%d %H:%M", "%Y-%m-%d %H:%M:%S"):
                            try:
                                fm = datetime.strptime(row["fecha_muestreo"], fmt)
                                break
                            except ValueError:
                                pass

                    if fm:
                        fecha_extraccion_estimada = (
                            fm + timedelta(days=int(tas_min))
                        ).strftime("%Y-%m-%d")

        except Exception as e:
            print("ERROR PANEL SILO:", s["numero_qr"], e)


        registros.append({
            **dict(s),
            "grado": grado,
            "factor": factor,
            "tas_min": tas_min,
            "fecha_extraccion_estimada": fecha_extraccion_estimada,
            "eventos": eventos_pendientes
        })

    conn.close()
    return render_template(
    "panel.html",
    registros=registros,
    puede_form=tiene_permiso("form"),
    puede_comercial=tiene_permiso("comercial"),
    puede_admin=tiene_permiso("admin"),
    puede_panel=tiene_permiso("panel")
)
    
# ======================
# COMERCIAL – PANTALLA
# ======================
@app.route("/comercial")
@login_required
def comercial():
    if not tiene_permiso("comercial"):
        return acceso_denegado("comercial")

    conn = get_db()

    rows = conn.execute("""
    SELECT cereal,
           pizarra_auto,
           fuente,
           fecha_fuente,
           pizarra_manual,
           usar_manual,
           obs_precio,
           dolar,
           fecha
    FROM mercado
    ORDER BY cereal
""").fetchall()

    dolar_info = conn.execute("""
        SELECT dolar, fecha
        FROM mercado
        WHERE dolar > 0
        ORDER BY fecha DESC
        LIMIT 1
    """).fetchone()

    conn.close()

    return render_template(
    "comercial.html",
    mercado=rows,
    dolar_info=dolar_info,
    puede_comparador=tiene_permiso("comparador")
)

# ======================
# COMPARADOR COMERCIAL
# ======================
@app.route("/comercial/<cereal>")
@login_required
def comparador(cereal):
    if not tiene_permiso("comparador"):
        return acceso_denegado("comparador")
    conn = get_db()

    rows = conn.execute("""
        SELECT
            s.numero_qr,

            -- FACTOR PROMEDIO
            (
              SELECT ROUND(AVG(a.factor),4)
              FROM analisis a
              JOIN muestreos m ON m.id = a.id_muestreo
              WHERE m.numero_qr = s.numero_qr
                AND a.factor IS NOT NULL
            ) AS factor_prom,

            -- HUMEDAD PROMEDIO
            (
              SELECT ROUND(AVG(a.humedad),2)
              FROM analisis a
              JOIN muestreos m ON m.id = a.id_muestreo
              WHERE m.numero_qr = s.numero_qr
                AND a.humedad IS NOT NULL
            ) AS humedad_prom,

            -- INSECTOS: si hay al menos uno en cualquier sección
            (
              SELECT COUNT(*)
              FROM analisis a
              JOIN muestreos m ON m.id = a.id_muestreo
              WHERE m.numero_qr = s.numero_qr
                AND a.insectos = 1
            ) AS tiene_insectos

        FROM silos s
        WHERE s.estado_silo = 'Activo'
          AND s.cereal = ?
        ORDER BY s.numero_qr
    """, (cereal,)).fetchall()

    conn.close()

    silos = [
        {
            **dict(r),
            "tiene_insectos": True if r["tiene_insectos"] > 0 else False
        }
        for r in rows
    ]

    return render_template(
        "comparador.html",
        cereal=cereal,
        silos=silos
    )

# ======================
# COMERCIAL – API
# ======================
@app.route("/api/mercado/manual", methods=["POST"])
@login_required
def mercado_manual():
    if not tiene_permiso("comercial"):
        return acceso_denegado("comercial")
    d = request.get_json()

    conn = get_db()
    conn.execute("""
        UPDATE mercado
        SET
            pizarra_manual=?,
            usar_manual=?,
            obs_precio=?,
            dolar=?,
            fecha=CURRENT_TIMESTAMP
        WHERE cereal=?
    """, (
        d.get("pizarra_manual"),
        1 if d.get("usar_manual") else 0,
        d.get("obs_precio"),
        d.get("dolar"),
        d["cereal"]
    ))
    conn.commit()
    conn.close()

    return jsonify(ok=True)

# ======================
# FORM
# ======================
@app.route("/form")
@login_required
def form():
    if not tiene_permiso("form"):
        return acceso_denegado("form")

    return render_template(
        "form.html",
        puede_calado=tiene_permiso("calado")
    )

# ======================
# RESTO DEL ARCHIVO
# ======================
# ⚠️ DESDE ACÁ NO SE TOCA NADA
# (tu código original sigue igual)

# ======================
# REGISTRAR SILO (FIX JSON)
# ======================
@app.route("/api/registrar_silo", methods=["POST"])
@login_required
def registrar_silo():
    if not tiene_permiso("form"):
        return acceso_denegado("form")
    d = request.get_json(force=True, silent=True)

    if not d or not d.get("numero_qr"):
        return jsonify(ok=False, error="QR faltante"), 400

    conn = get_db()
    conn.execute("""
        INSERT INTO silos (
            numero_qr,
            cereal,
            estado_grano,
            estado_silo,
            metros,
            lat,
            lon,
            fecha_confeccion
        ) VALUES (?,?,?,?,?,?,?,?)
    """, (
        d.get("numero_qr"),
        d.get("cereal"),
        d.get("estado_grano"),
        "Activo",
        d.get("metros"),
        d.get("lat"),
        d.get("lon"),
        ahora().strftime("%Y-%m-%d %H:%M")
    ))

    conn.commit()
    conn.close()
    return jsonify(ok=True)

# ======================
# API — NUEVO MUESTREO
# ======================
@app.route("/api/nuevo_muestreo", methods=["POST"])
@login_required
def api_nuevo_muestreo():
    if not tiene_permiso("calado"):
        return redirect(url_for("no_autorizado", seccion="calado"))
    d = request.get_json(force=True, silent=True) or {}
    qr = d.get("qr")

    if not qr:
        return jsonify(error="QR faltante"), 400

    conn = get_db()

    # 🔒 BLOQUEO SI ESTÁ EXTRAÍDO
    silo = conn.execute(
        "SELECT estado_silo FROM silos WHERE numero_qr=?",
        (qr,)
    ).fetchone()

    if not silo or silo["estado_silo"] == "Extraído":
        conn.close()
        return jsonify(
            ok=False,
            error="El silo ya fue extraído. No se pueden cargar nuevos muestreos."
        ), 400

    # ✅ SI ESTÁ ACTIVO, CREA EL MUESTREO
    cur = conn.cursor()
    cur.execute("""
        INSERT INTO muestreos (numero_qr, fecha_muestreo)
        VALUES (?,?)
    """, (qr, ahora().strftime("%Y-%m-%d %H:%M")))

    conn.commit()
    mid = cur.lastrowid
    conn.close()

    return jsonify(ok=True, id_muestreo=mid)
# ======================
# INFORMAR CALADO (DESDE FORM)
# ======================
@app.route("/api/informar_calado", methods=["POST"])
@login_required
def informar_calado():
    if not tiene_permiso("calado"):
        return acceso_denegado("calado")
    d = request.get_json(force=True, silent=True) or {}
    qr = d.get("numero_qr")

    if not qr:
        return jsonify(ok=False, error="QR faltante"), 400

    conn = get_db()

    # 🔒 validar silo
    silo = conn.execute(
        "SELECT estado_silo FROM silos WHERE numero_qr=?",
        (qr,)
    ).fetchone()

    if not silo:
        conn.close()
        return jsonify(ok=False, error="Silo inexistente"), 400

    if silo["estado_silo"] == "Extraído":
        conn.close()
        return jsonify(
            ok=False,
            error="El silo ya fue extraído. No se puede registrar calado."
        ), 400

    # ✅ crear muestreo
    cur = conn.cursor()
    cur.execute("""
        INSERT INTO muestreos (numero_qr, fecha_muestreo)
        VALUES (?,?)
    """, (qr, ahora().strftime("%Y-%m-%d %H:%M")))

    id_muestreo = cur.lastrowid

    # 🧪 ¿informó temperatura?
    if d.get("informar_temperatura"):
        for seccion, campo in [
            ("punta", "temp_punta"),
            ("medio", "temp_medio"),
            ("final", "temp_final")
        ]:
            temp = d.get(campo)

            if temp not in (None, "", ""):
                try:
                    temp = float(temp)
                except ValueError:
                    temp = None

            cur.execute("""
                INSERT INTO analisis (
                    id_muestreo, seccion, temperatura
                ) VALUES (?,?,?)
            """, (
                id_muestreo,
                seccion,
                temp
            ))

    conn.commit()
    conn.close()

    return jsonify(ok=True, id_muestreo=id_muestreo)

# ======================
# ANALISIS — SECCION
# ======================
@app.route("/api/analisis_seccion", methods=["POST"])
@login_required
def guardar_analisis_seccion():
    if not tiene_permiso("laboratorio"):
        return acceso_denegado("calado")
    d = request.get_json(force=True, silent=True) or {}

    def to_float(x):
        try:
            return float(x)
        except (TypeError, ValueError):
            return None

    for k in ["temperatura","humedad","ph","danados","quebrados","materia_extrana","olor","moho","chamico"]:
        d[k] = to_float(d.get(k))

    d["insectos"] = 1 if d.get("insectos") else 0

    conn = get_db()
    cur = conn.cursor()

    existente = cur.execute("""
        SELECT id FROM analisis
        WHERE id_muestreo=? AND seccion=?
    """, (d["id_muestreo"], d["seccion"])).fetchone()

    res = calcular_comercial(d["cereal"], d)

    valores = (
        d["id_muestreo"],
        d["seccion"],
        d["temperatura"],
        d["humedad"],
        d["ph"],
        d["danados"],
        d["quebrados"],
        d["materia_extrana"],
        d["olor"],
        d["moho"],
        d["insectos"],
        d["chamico"],
        res["grado"],
        res["factor"],
        res["tas"]
    )

    if existente:
        cur.execute("""
            UPDATE analisis SET
                temperatura=?, humedad=?, ph=?,
                danados=?, quebrados=?, materia_extrana=?,
                olor=?, moho=?, insectos=?, chamico=?,
                grado=?, factor=?, tas=?
            WHERE id_muestreo=? AND seccion=?
        """, valores[2:] + valores[:2])
    else:
        cur.execute("""
            INSERT INTO analisis (
                id_muestreo, seccion,
                temperatura, humedad, ph,
                danados, quebrados, materia_extrana,
                olor, moho, insectos, chamico,
                grado, factor, tas
            ) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
        """, valores)

    conn.commit()
    conn.close()
    return jsonify(ok=True)

# ======================
# MONITOREO
# ======================
@app.route("/api/monitoreo", methods=["POST"])
@login_required
def nuevo_monitoreo():
    if not tiene_permiso("form"):
        return acceso_denegado("form")
    qr = request.form.get("numero_qr")
    tipo = request.form.get("tipo")
    detalle = request.form.get("detalle")
    foto_evento = request.files.get("foto")

    conn = get_db()

    silo = conn.execute(
        "SELECT estado_silo FROM silos WHERE numero_qr=?",
        (qr,)
    ).fetchone()

    if not silo or silo["estado_silo"] == "Extraído":
        conn.close()
        return jsonify(
            ok=False,
            error="El silo está extraído. No se pueden cargar eventos."
        ), 400

    path_evento = None
    if foto_evento:
        os.makedirs("static/monitoreos", exist_ok=True)
        path_evento = f"static/monitoreos/{datetime.now().timestamp()}_{foto_evento.filename}"
        foto_evento.save(path_evento)

    conn.execute("""
        INSERT INTO monitoreos (
            numero_qr, fecha_evento, tipo, detalle, foto_evento
        ) VALUES (?,?,?,?,?)
    """, (
        qr,
        ahora().strftime("%Y-%m-%d %H:%M"),
        tipo,
        detalle,
        path_evento
    ))

    conn.commit()
    conn.close()
    return jsonify(ok=True)
    
# ======================
# MONITOREO PENDIENTE
# ======================
@app.route("/api/monitoreo/pendiente/<qr>")
@login_required
def monitoreos_pendientes(qr):
    conn = get_db()
    rows = conn.execute("""
        SELECT id, tipo, fecha_evento
        FROM monitoreos
        WHERE numero_qr = ?
          AND resuelto = 0
        ORDER BY fecha_evento DESC
    """, (qr,)).fetchall()
    conn.close()

    return jsonify([
        {
            "id": r["id"],
            "tipo": r["tipo"],
            "fecha": r["fecha_evento"]
        } for r in rows
    
    ])
    
# ======================
# RESOLVER MONITOREO
# ======================
@app.route("/api/monitoreo/resolver", methods=["POST"])
@login_required
def resolver_monitoreo():
    id_monitoreo = request.form.get("id_monitoreo")
    foto = request.files.get("foto")

    if not id_monitoreo:
        return jsonify(ok=False, error="ID faltante"), 400

    path_resolucion = None
    if foto:
        os.makedirs("static/monitoreos", exist_ok=True)
        path_resolucion = f"static/monitoreos/resuelto_{datetime.now().timestamp()}_{foto.filename}"
        foto.save(path_resolucion)

    conn = get_db()
    conn.execute("""
        UPDATE monitoreos SET
            resuelto = 1,
            fecha_resolucion = ?,
            foto_resolucion = ?
        WHERE id = ?
    """, (
        ahora().strftime("%Y-%m-%d %H:%M"),
        path_resolucion,
        id_monitoreo
    ))
    conn.commit()
    conn.close()

    return jsonify(ok=True)
    
# ======================
# MONITOREOS RESUELTOS
# ======================
@app.route("/api/monitoreo/resueltos/<qr>")
@login_required
def monitoreos_resueltos(qr):
    conn = get_db()
    rows = conn.execute("""
        SELECT tipo, fecha_resolucion
        FROM monitoreos
        WHERE numero_qr = ?
          AND resuelto = 1
        ORDER BY fecha_resolucion DESC
    """, (qr,)).fetchall()
    conn.close()

    return jsonify([
        {
            "tipo": r["tipo"],
            "fecha": r["fecha_resolucion"]
        } for r in rows
    ])
# ======================
# EXTRACCION
# ======================
@app.route("/api/extraccion", methods=["POST"])
@login_required
def registrar_extraccion():
    if not tiene_permiso("form"):
        return acceso_denegado("form")
    d = request.get_json(force=True, silent=True)

    conn = get_db()
    conn.execute("""
        UPDATE silos SET
            estado_silo=?,
            fecha_extraccion=?
        WHERE numero_qr=?
    """, (
        d["estado_silo"],
        ahora().strftime("%Y-%m-%d %H:%M"),
        d["numero_qr"]
    ))

    conn.commit()
    conn.close()
    return jsonify(ok=True)
# ======================
# SILO (DETALLE)
# ======================
@app.route("/silo/<qr>")
@login_required
def ver_silo(qr):
    if not tiene_permiso("panel"):
        return acceso_denegado("panel")
    conn = get_db()

    silo = conn.execute(
        "SELECT * FROM silos WHERE numero_qr=?",
        (qr,)
    ).fetchone()

    if not silo:
        conn.close()
        return "Silo no encontrado", 404

    mercado = conn.execute("""
        SELECT
            CASE
                WHEN usar_manual = 1 THEN pizarra_manual
                ELSE pizarra_auto
            END AS pizarra,
            dolar
        FROM mercado
        WHERE cereal = ?
    """, (silo["cereal"],)).fetchone()

    muestreos_raw = conn.execute("""
        SELECT m.id, m.fecha_muestreo,
               CAST(julianday('now') - julianday(m.fecha_muestreo) AS INT) dias
        FROM muestreos m
        WHERE m.numero_qr=?
        ORDER BY m.fecha_muestreo DESC
    """, (qr,)).fetchall()

    muestreos = []
    precio_estimado = None
    precio_usd = None
    factor_prom = None
    tas_usada = None
    analisis_pendiente = False

    for idx, m in enumerate(muestreos_raw):
        analisis = conn.execute("""
            SELECT seccion, grado, factor, tas, temperatura
            FROM analisis
            WHERE id_muestreo=?
        """, (m["id"],)).fetchall()

        por_seccion = {a["seccion"]: a for a in analisis}

        if idx == 0:
            if not analisis:
                analisis_pendiente = True
            else:
                factores = []
                tass = []

                for sec in ["punta", "medio", "final"]:
                    a = por_seccion.get(sec)
                    if a:
                        if a["factor"] is not None:
                            factores.append(a["factor"])
                        if a["tas"] is not None:
                            tass.append(a["tas"])

                if factores:
                    factor_prom = round(sum(factores) / len(factores), 4)

                if tass:
                    tas_usada = min(tass)

                if mercado and factor_prom and mercado["pizarra"] and mercado["dolar"]:
                    precio_estimado = round(mercado["pizarra"] * factor_prom, 2)
                    precio_usd = round(precio_estimado / mercado["dolar"], 2)

        muestreos.append({
            "id": m["id"],
            "fecha_muestreo": m["fecha_muestreo"],
            "dias": m["dias"],
            "punta": por_seccion.get("punta"),
            "medio": por_seccion.get("medio"),
            "final": por_seccion.get("final")
        })

    eventos_pendientes = conn.execute("""
        SELECT tipo, fecha_evento, foto_evento
        FROM monitoreos
        WHERE numero_qr = ?
          AND resuelto = 0
        ORDER BY fecha_evento DESC
    """, (qr,)).fetchall()

    eventos_resueltos = conn.execute("""
        SELECT tipo, fecha_resolucion, foto_resolucion
        FROM monitoreos
        WHERE numero_qr = ?
          AND resuelto = 1
        ORDER BY fecha_resolucion DESC
    """, (qr,)).fetchall()

    conn.close()

    return render_template(
    "silo.html",
    silo=silo,
    muestreos=muestreos,
    eventos_pendientes=eventos_pendientes,
    eventos_resueltos=eventos_resueltos,
    mercado=mercado,
    precio_estimado=precio_estimado,
    precio_usd=precio_usd,
    factor_prom=factor_prom,
    tas_usada=tas_usada,
    analisis_pendiente=analisis_pendiente,
    puede_calado=tiene_permiso("calado")   # 👈 agregar
)

# ======================
# VER MUESTREO
# ======================
@app.route("/muestreo/<int:id>")
@login_required
def ver_muestreo(id):
    if not tiene_permiso("panel"):
        return redirect(url_for("no_autorizado", seccion="panel"))
    conn = get_db()

    muestreo = conn.execute("""
        SELECT m.*, s.numero_qr, s.cereal
        FROM muestreos m
        JOIN silos s ON s.numero_qr = m.numero_qr
        WHERE m.id=?
    """, (id,)).fetchone()

    if not muestreo:
        conn.close()
        return "Muestreo no encontrado", 404

    analisis = conn.execute("""
        SELECT *
        FROM analisis
        WHERE id_muestreo=?
        ORDER BY seccion
    """, (id,)).fetchall()

    conn.close()

    return render_template(
    "muestreo.html",
    muestreo=muestreo,
    analisis=analisis,
    puede_laboratorio=tiene_permiso("laboratorio")   # 👈 agregar
)
# ======================
# EXPORT EXCEL ORDENADO
# ======================
@app.route("/api/export")
@login_required
def exportar_excel():
    if not tiene_permiso("panel"):
        return acceso_denegado("panel")
    from openpyxl import Workbook
    from openpyxl.styles import Font
    from datetime import datetime, timedelta

    conn = get_db()

    silos = conn.execute("""
        SELECT s.*,
        (
            SELECT m.id
            FROM muestreos m
            WHERE m.numero_qr = s.numero_qr
            ORDER BY m.id DESC
            LIMIT 1
        ) ultimo_muestreo
        FROM silos s
        ORDER BY s.cereal, s.numero_qr
    """).fetchall()

    wb = Workbook()
    wb.remove(wb.active)

    cereales = ["Soja", "Maíz", "Trigo", "Girasol"]

    for cereal in cereales:
        ws = wb.create_sheet(title=cereal)

        ws["A1"] = "PRECIO BASE ($)"
        ws["A1"].font = Font(bold=True)
        ws["B1"] = 0

        headers = [
            "QR",
            "Fecha confección",
            "Estado silo",
            "Estado grano",
            "Metros",
            "Grado",
            "Factor",
            "TAS mín",
            "Fecha extracción estimada",
            "Precio estimado"
        ]

        ws.append([])
        ws.append(headers)

        for col in range(1, len(headers) + 1):
            ws.cell(row=3, column=col).font = Font(bold=True)

        row_excel = 4

        for s in silos:

            if s["cereal"] != cereal:
                continue

            grado = None
            factor = None
            tas_min = None
            fecha_estimada = None

            # ==============================
            # ANALISIS
            # ==============================
            if s["ultimo_muestreo"]:

                analisis = conn.execute("""
                    SELECT grado, factor, tas
                    FROM analisis
                    WHERE id_muestreo=?
                """, (s["ultimo_muestreo"],)).fetchall()

                grados = []
                factores = []
                tass = []

                for a in analisis:

                    try:
                        if a["grado"] is not None:
                            grados.append(int(a["grado"]))
                    except:
                        pass

                    try:
                        if a["factor"] is not None:
                            factores.append(float(a["factor"]))
                    except:
                        pass

                    try:
                        if a["tas"] is not None:
                            tass.append(int(a["tas"]))
                    except:
                        pass

                if grados:
                    grado = max(grados)

                if factores:
                    factor = round(sum(factores) / len(factores), 4)

                if tass:
                    tas_min = min(tass)

                    row_fecha = conn.execute("""
                        SELECT fecha_muestreo
                        FROM muestreos
                        WHERE id=?
                    """, (s["ultimo_muestreo"],)).fetchone()

                    if row_fecha and row_fecha["fecha_muestreo"]:
                        try:
                            fm = datetime.strptime(
                                row_fecha["fecha_muestreo"],
                                "%Y-%m-%d %H:%M"
                            )
                            fecha_estimada = fm + timedelta(days=int(tas_min))
                        except:
                            pass

            # ==============================
            # ESCRIBIR FILA EXCEL
            # ==============================
            ws.cell(row=row_excel, column=1, value=s["numero_qr"])

            if s["fecha_confeccion"]:
                try:
                    fecha_obj = datetime.strptime(
                        s["fecha_confeccion"],
                        "%Y-%m-%d %H:%M"
                    )
                    ws.cell(row=row_excel, column=2, value=fecha_obj)
                except:
                    pass

            ws.cell(row=row_excel, column=3, value=s["estado_silo"])
            ws.cell(row=row_excel, column=4, value=s["estado_grano"])
            ws.cell(row=row_excel, column=5, value=s["metros"])
            ws.cell(row=row_excel, column=6, value=grado)
            ws.cell(row=row_excel, column=7, value=factor)
            ws.cell(row=row_excel, column=8, value=tas_min)
            ws.cell(row=row_excel, column=9, value=fecha_estimada)

            ws.cell(
                row=row_excel,
                column=10,
                value=f"=B1*G{row_excel}"
            )

            row_excel += 1

    conn.close()

    filename = "silos_comercial.xlsx"
    wb.save(filename)

    return send_file(
        filename,
        as_attachment=True,
        download_name=filename
    )

# ======================
# DÓLAR OFICIAL
# ======================
def obtener_dolar_oficial():
    try:
        r = requests.get(
            "https://api.bluelytics.com.ar/v2/latest",
            timeout=10
        )
        data = r.json()
        return data["oficial"]["value_avg"]
    except Exception as e:
        print("Error dólar:", e)
        return None

def obtener_pizarra_auto(cereal):
    try:
        # normalizar
        c = cereal.strip().lower()

        precios_rosario = {
            "soja": 275000,
            "maiz": 160000,
            "trigo": 185000,
            "girasol": 295000
        }

        precio = precios_rosario.get(c)

        if precio is None:
            print("⚠️ Cereal no encontrado:", cereal)
            return None

        return {
            "precio": precio,
            "fuente": "Rosario (CAC) – ACAbase",
            "fecha": ahora().strftime("%Y-%m-%d %H:%M")
        }

    except Exception as e:
        print("Error pizarra:", e)
        return None

@app.route("/api/actualizar_dolar", methods=["POST"])
@login_required
def actualizar_dolar():
    if not tiene_permiso("comercial"):
        return acceso_denegado("comercial")

    dolar = obtener_dolar_oficial()
    if dolar is None:
        return jsonify({"ok": False, "error": "No se pudo obtener el dólar"})

    conn = get_db()
    conn.execute("""
        UPDATE mercado
        SET dolar = ?, fecha = datetime('now')
    """, (dolar,))
    conn.commit()
    conn.close()

    return jsonify({"ok": True, "dolar": dolar})

from bs4 import BeautifulSoup
import re

def obtener_pizarra_auto(cereal):
    precios_mock = {
        "Soja": 275000,
        "Maíz": 160000,
        "Trigo": 185000,
        "Girasol": 295000
    }

    precio = precios_mock.get(cereal)
    if not precio:
        return None

    return {
        "precio": precio,
        "fuente": "Carga manual / referencia CAC",
        "fecha": ahora().strftime("%Y-%m-%d %H:%M")
    }

@app.route("/api/actualizar_pizarra", methods=["POST"])
@login_required
def actualizar_pizarra():
    conn = get_db()
    rows = conn.execute("SELECT cereal FROM mercado").fetchall()

    for r in rows:
        cereal = r["cereal"]
        data = obtener_pizarra_auto(cereal)
        if not data:
            continue

        conn.execute("""
            UPDATE mercado
            SET pizarra_auto = ?,
                fuente = ?,
                fecha_fuente = ?,
                fecha = datetime('now')
            WHERE cereal = ?
        """, (
            data["precio"],
            data["fuente"],
            data["fecha"],
            cereal
        ))

    conn.commit()
    conn.close()
    return jsonify(ok=True)

# ======================
# RUN
# ======================
if __name__ == "__main__":
    app.run(debug=True)
