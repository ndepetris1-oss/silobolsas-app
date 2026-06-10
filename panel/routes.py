from flask import Blueprint, render_template, session, redirect, url_for, send_file
from utils.auditoria import registrar_auditoria
from flask_login import login_required, current_user
from db import get_db
from permissions import tiene_permiso, acceso_denegado
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font
from io import BytesIO

panel_bp = Blueprint("panel", __name__)


# =====================================================
# Compatibilidad SQLite / PostgreSQL
# =====================================================
def _is_postgres_backend(conn):
    for obj in (getattr(conn, "cursor", None), conn):
        if obj is not None:
            mod = getattr(obj.__class__, "__module__", "").lower()
            if "psycopg2" in mod:
                return True
    return False


def db_execute(conn, query, params=None):
    """Execute portable: convierte ? -> %s en PostgreSQL."""
    if _is_postgres_backend(conn) and "?" in query:
        query = query.replace("?", "%s")
    _real_exec = object.__getattribute__(conn, "execute")
    try:
        return _real_exec(query, params or ())
    except Exception:
        _rb = getattr(conn, "rollback", None)
        if callable(_rb):
            try:
                _rb()
            except Exception:
                pass
        raise


# KG estimados por metro lineal de silo bolsa según cereal
KG_POR_METRO = {
    "Soja": 3100,
}
KG_POR_METRO_DEFAULT = 1000



def empresa_actual():
    if current_user.es_superadmin:
        return session.get("empresa_contexto")
    return current_user.empresa_id


@panel_bp.route("/silo/<qr>")
@login_required
def ver_silo(qr):

    if not tiene_permiso("panel") and not tiene_permiso("laboratorio"):
        return acceso_denegado("panel")

    conn = get_db()
    empresa_id = empresa_actual()

    silo = db_execute(conn, """
        SELECT * FROM silos
        WHERE numero_qr=? AND empresa_id=?
    """, (qr, empresa_id)).fetchone()

    if not silo:
        conn.close()
        return "Silo no encontrado", 404

    mercado = db_execute(conn, """
        SELECT
            CASE
                WHEN usar_manual = 1 THEN pizarra_manual
                ELSE pizarra_auto
            END AS pizarra,
            dolar
        FROM mercado
        WHERE cereal = ? AND empresa_id = ?
    """, (silo["cereal"], empresa_id)).fetchone()

    muestreos_raw_db = db_execute(conn, """
        SELECT m.id, m.fecha_muestreo
        FROM muestreos m
        WHERE m.numero_qr=? AND empresa_id=?
        ORDER BY m.fecha_muestreo DESC
    """, (qr, empresa_id)).fetchall()

    # ─── Helper para parsear fechas de ambas fuentes ───────────────────────
    def _parsear_fecha(val):
        if val is None:
            return None
        if not isinstance(val, str):
            return val
        for fmt in ("%Y-%m-%d %H:%M", "%Y-%m-%d %H:%M:%S", "%Y-%m-%d"):
            try:
                return datetime.strptime(val, fmt)
            except ValueError:
                continue
        return None

    muestreos_raw = []
    for m in muestreos_raw_db:
        fecha_val = m["fecha_muestreo"]
        dias = None
        if fecha_val:
            try:
                fecha = _parsear_fecha(fecha_val)
                if fecha:
                    dias = (datetime.now() - fecha).days
            except:
                dias = None
        muestreos_raw.append({"id": m["id"], "fecha_muestreo": fecha_val, "dias": dias})

    muestreos = []
    precio_estimado = None
    precio_usd = None
    factor_prom = None
    tas_usada = None
    analisis_pendiente = False

    # Variables del calado más reciente con análisis
    factor_calado = None
    tas_calado = None
    fecha_calado = None

    for idx, m in enumerate(muestreos_raw):

        analisis = db_execute(conn, """
            SELECT seccion, grado, factor, tas, temperatura
            FROM analisis
            WHERE id_muestreo=? AND empresa_id=?
        """, (m["id"], empresa_id)).fetchall()

        por_seccion = {a["seccion"]: a for a in analisis}

        if idx == 0:
            if not analisis:
                analisis_pendiente = True
            else:
                factores = []
                tass = []

                for sec in ["punta", "medio", "final"]:
                    a = por_seccion.get(sec)
                    if a:
                        if a["factor"] is not None:
                            factores.append(a["factor"])
                        if a["tas"] is not None:
                            tass.append(a["tas"])

                if factores:
                    factor_calado = round(sum(factores) / len(factores), 4)

                if tass:
                    tas_calado = min(tass)

                fecha_calado = _parsear_fecha(m["fecha_muestreo"])

        muestreos.append({
            "id": m["id"],
            "fecha_muestreo": m["fecha_muestreo"],
            "dias": m["dias"],
            "punta": por_seccion.get("punta"),
            "medio": por_seccion.get("medio"),
            "final": por_seccion.get("final")
        })

    eventos_pendientes = db_execute(conn, """
        SELECT tipo, fecha_evento, foto_evento
        FROM monitoreos
        WHERE numero_qr = ? AND empresa_id=? AND resuelto = 0
        ORDER BY fecha_evento DESC
    """, (qr, empresa_id)).fetchall()

    eventos_resueltos = db_execute(conn, """
        SELECT tipo, fecha_resolucion, foto_resolucion
        FROM monitoreos
        WHERE numero_qr = ? AND empresa_id=? AND resuelto = 1
        ORDER BY fecha_resolucion DESC
    """, (qr, empresa_id)).fetchall()

    try:
        cargas_raw = db_execute(conn, """
            SELECT id, fecha, kg, temperatura, humedad, danados,
                quebrados, materia_extrana, olor, moho, insectos,
                chamico, grado, factor, tas
            FROM llenado
            WHERE numero_qr=? AND empresa_id=?
            ORDER BY fecha DESC
        """, (qr, empresa_id)).fetchall()

        cargas_llenado = [dict(x) for x in cargas_raw]
        kg_total = sum(float(c["kg"] or 0) for c in cargas_llenado)

    except Exception:
        conn.rollback()
        cargas_llenado = []
        kg_total = 0

    # ─── Calcular factor/tas del llenado más reciente ──────────────────────
    factor_llenado = None
    tas_llenado = None
    fecha_llenado = None

    if cargas_llenado:
        # fecha de la carga más reciente (ya ordenadas DESC)
        fecha_llenado = _parsear_fecha(cargas_llenado[0].get("fecha"))

        # factor ponderado por kg de TODAS las cargas
        kg_sum = 0.0
        factor_sum = 0.0
        tass_ll = []
        for c in cargas_llenado:
            f = c.get("factor")
            kg = float(c.get("kg") or 0)
            if f is not None and kg > 0:
                factor_sum += float(f) * kg
                kg_sum += kg
            t = c.get("tas")
            if t is not None:
                tass_ll.append(int(t))
        if kg_sum > 0:
            factor_llenado = round(factor_sum / kg_sum, 4)
        if tass_ll:
            tas_llenado = min(tass_ll)

    # ─── Elegir la fuente más reciente ─────────────────────────────────────
    usar_llenado = False
    if factor_llenado is not None:
        if fecha_llenado and fecha_calado:
            # ambas fuentes tienen datos: usar la más reciente
            if fecha_llenado > fecha_calado:
                usar_llenado = True
        elif fecha_llenado and not fecha_calado:
            # solo hay llenado
            usar_llenado = True

    if usar_llenado:
        factor_prom = factor_llenado
        tas_usada = tas_llenado
        fuente_precio = "llenado"
    else:
        factor_prom = factor_calado
        tas_usada = tas_calado
        fuente_precio = "calado"

    # ─── Calcular precio estimado con la fuente elegida ────────────────────
    if mercado and factor_prom and mercado["pizarra"] and mercado["dolar"]:
        precio_estimado = round(mercado["pizarra"] * factor_prom, 2)
        precio_usd = round(precio_estimado / mercado["dolar"], 2)

    # ─── Camionadas de vaciado ────────────────────────────────────────────────
    camionadas = []
    kg_extraidos = 0
    comparativo = None

    if silo["estado_silo"] in ("En extracción", "Extraído"):
        try:
            cam_raw = db_execute(conn, """
                SELECT *
                FROM vaciado
                WHERE numero_qr=? AND empresa_id=?
                ORDER BY nro_camion ASC
            """, (qr, empresa_id)).fetchall()
            camionadas = [dict(c) for c in cam_raw]
            kg_extraidos = sum(float(c["kg"] or 0) for c in camionadas if c.get("kg"))
        except Exception:
            try:
                conn.rollback()
            except:
                pass
            camionadas = []
            kg_extraidos = 0

    if silo["estado_silo"] == "Extraído" and camionadas:
        # Factor ponderado del vaciado
        kg_sum_vac = sum(float(c["kg"] or 0) for c in camionadas if c.get("factor") and c.get("kg"))
        factor_vac = None
        if kg_sum_vac > 0:
            factor_vac = round(
                sum(float(c["factor"]) * float(c["kg"] or 0) for c in camionadas if c.get("factor") and c.get("kg"))
                / kg_sum_vac, 4
            )
        dif_kg = round(kg_extraidos - kg_total, 0) if kg_total > 0 else None
        dif_factor = round((factor_vac - factor_prom) * 100, 3) if (factor_vac and factor_prom) else None
        comparativo = dict(
            kg_llenado=kg_total,
            kg_vaciado=kg_extraidos,
            dif_kg=dif_kg,
            factor_llenado=factor_prom,
            factor_vaciado=factor_vac,
            dif_factor=dif_factor,
            fecha_extraccion=silo["fecha_extraccion"],
        )

    conn.close()

    return render_template(
        "silo.html",
        silo=silo,
        muestreos=muestreos,
        eventos_pendientes=eventos_pendientes,
        eventos_resueltos=eventos_resueltos,
        mercado=mercado,
        precio_estimado=precio_estimado,
        precio_usd=precio_usd,
        factor_prom=factor_prom,
        tas_usada=tas_usada,
        analisis_pendiente=analisis_pendiente,
        puede_calado=tiene_permiso("calado"),
        puede_comercial=tiene_permiso("comercial"),
        puede_admin=tiene_permiso("admin"),
        dif_matba=None,
        cargas_llenado=cargas_llenado,
        kg_total=kg_total,
        fuente_precio=fuente_precio,
        camionadas=camionadas,
        kg_extraidos=kg_extraidos,
        comparativo=comparativo,
    )


@panel_bp.route("/muestreo/<int:id>")
@login_required
def ver_muestreo(id):

    if not tiene_permiso("panel"):
        return acceso_denegado("panel")

    conn = get_db()
    empresa_id = empresa_actual()

    muestreo = db_execute(conn, """
        SELECT m.*, s.numero_qr, s.cereal
        FROM muestreos m
        JOIN silos s
        ON s.numero_qr = m.numero_qr
        AND s.empresa_id = m.empresa_id
        WHERE m.id=? AND m.empresa_id=?
    """, (id, empresa_id)).fetchone()

    if not muestreo:
        conn.close()
        return "Muestreo no encontrado", 404

    analisis = db_execute(conn, """
        SELECT *
        FROM analisis
        WHERE id_muestreo=? AND empresa_id=?
        ORDER BY seccion
    """, (id, empresa_id)).fetchall()

    conn.close()

    return render_template(
        "muestreo.html",
        muestreo=muestreo,
        analisis=analisis,
        puede_laboratorio=tiene_permiso("laboratorio")
    )


@panel_bp.route("/")
@panel_bp.route("/panel")
@login_required
def panel():

    if not tiene_permiso("panel"):
        return acceso_denegado("panel")

    conn = get_db()

    if current_user.es_superadmin:

        if "empresa_contexto" not in session:

            empresas = db_execute(conn, """
                SELECT id, nombre
                FROM empresas
                WHERE activa = 1
                ORDER BY nombre
            """).fetchall()

            conn.close()

            return render_template(
                "seleccionar_empresa.html",
                empresas=empresas
            )

        empresa_id = session["empresa_contexto"]

    else:
        empresa_id = current_user.empresa_id

    silos = db_execute(conn, """
        SELECT *
        FROM silos
        WHERE empresa_id=?
        ORDER BY fecha_confeccion DESC
    """, (empresa_id,)).fetchall()

    registros = []

    for s in silos:

        ultimo = db_execute(conn, """
            SELECT id, fecha_muestreo
            FROM muestreos
            WHERE numero_qr=? AND empresa_id=?
            ORDER BY fecha_muestreo DESC
            LIMIT 1
        """, (s["numero_qr"], empresa_id)).fetchone()

        grado = None
        factor_prom = None
        tas_min = None
        fecha_estimada = None

        if ultimo:

            analisis = db_execute(conn, """
                SELECT grado, factor, tas
                FROM analisis
                WHERE id_muestreo=? AND empresa_id=?
            """, (ultimo["id"], empresa_id)).fetchall()

            grados = []
            factores = []
            tass = []

            for a in analisis:

                g = a["grado"]

                if g is not None:
                    if str(g).upper() == "F/E":
                        grados = ["F/E"]
                    else:
                        try:
                            grados.append(int(g))
                        except:
                            pass

                if a["factor"] is not None:
                    try:
                        factores.append(float(a["factor"]))
                    except:
                        pass

                if a["tas"] is not None:
                    try:
                        tass.append(int(a["tas"]))
                    except:
                        pass

            if "F/E" in grados:
                grado = "F/E"
            elif grados:
                grado = max(grados)

            if factores:
                factor_prom = round(sum(factores) / len(factores), 4)

            if tass:
                tas_min = min(tass)

            if tas_min is not None and ultimo["fecha_muestreo"]:

                fecha_str = ultimo["fecha_muestreo"]
                fecha_base = None

                for fmt in ("%Y-%m-%d %H:%M", "%Y-%m-%d %H:%M:%S"):
                    try:
                        fecha_base = datetime.strptime(fecha_str, fmt)
                        break
                    except ValueError:
                        continue

                if fecha_base:
                    fecha_estimada = (
                        fecha_base + timedelta(days=tas_min)
                    ).strftime("%Y-%m-%d")

        eventos = db_execute(conn, """
            SELECT COUNT(*) as cant
            FROM monitoreos
            WHERE numero_qr=? AND empresa_id=? AND resuelto=0
        """, (s["numero_qr"], empresa_id)).fetchone()["cant"]

        try:
            kg_row = db_execute(conn, """
                SELECT COALESCE(SUM(kg), 0) as total
                FROM llenado
                WHERE numero_qr=? AND empresa_id=?
            """, (s["numero_qr"], empresa_id)).fetchone()
            kg_total = int(kg_row["total"]) if kg_row else 0
        except Exception:
            conn.rollback()
            kg_total = 0

        # Si no hay calado, usar datos ponderados del llenado
        fuente = "calado"
        if grado is None and factor_prom is None and tas_min is None:
            try:
                cargas = db_execute(conn, """
                    SELECT kg, factor, tas, fecha
                    FROM llenado
                    WHERE numero_qr=? AND empresa_id=?
                    ORDER BY fecha DESC
                """, (s["numero_qr"], empresa_id)).fetchall()

                if cargas:
                    fuente = "llenado"
                    cargas_con_factor = [c for c in cargas if c["factor"] is not None]
                    if cargas_con_factor:
                        kg_total_pond = sum(float(c["kg"] or 0) for c in cargas_con_factor)
                        if kg_total_pond > 0:
                            factor_pond = sum(
                                float(c["factor"]) * float(c["kg"] or 0)
                                for c in cargas_con_factor
                            ) / kg_total_pond
                            factor_prom = round(factor_pond, 4)
                        else:
                            factores = [float(c["factor"]) for c in cargas_con_factor]
                            factor_prom = round(sum(factores) / len(factores), 4)

                    tas_vals = [int(c["tas"]) for c in cargas if c["tas"] is not None]
                    if tas_vals:
                        tas_min = min(tas_vals)

                    if tas_min and cargas[0]["fecha"]:
                        fecha_str = cargas[0]["fecha"]
                        try:
                            for fmt in ("%Y-%m-%d %H:%M", "%Y-%m-%d %H:%M:%S", "%Y-%m-%d"):
                                try:
                                    fecha_base = datetime.strptime(fecha_str, fmt)
                                    break
                                except ValueError:
                                    continue
                            fecha_estimada = (
                                fecha_base + timedelta(days=tas_min)
                            ).strftime("%Y-%m-%d")
                        except:
                            pass
            except Exception:
                pass

        # kg extraídos en vaciado (solo para silos Extraídos o En extracción)
        kg_vaciado = 0
        if s["estado_silo"] in ("Extraído", "En extracción"):
            try:
                row_vac = conn.execute(
                    "SELECT COALESCE(SUM(kg),0) AS total FROM vaciado WHERE numero_qr=? AND empresa_id=?",
                    (s["numero_qr"], empresa_id)
                ).fetchone()
                kg_vaciado = int(row_vac["total"] or 0)
            except Exception:
                try: conn.rollback()
                except: pass

        registros.append({
            **dict(s),
            "grado": grado,
            "factor": factor_prom,
            "tas_min": tas_min,
            "fecha_extraccion_estimada": fecha_estimada,
            "eventos": eventos,
            "kg_total": kg_total,
            "kg_vaciado": kg_vaciado,
            "dif_kg": (kg_vaciado - kg_total) if s["estado_silo"] == "Extraído" and kg_total > 0 else None,
            "fuente_calidad": fuente
        })

    conn.close()

    total_activos       = sum(1 for r in registros if r.get("estado_silo") not in ("Extraído", "En extracción"))
    total_en_extraccion = sum(1 for r in registros if r.get("estado_silo") == "En extracción")
    total_extraidos     = sum(1 for r in registros if r.get("estado_silo") == "Extraído")
    con_alertas     = sum(1 for r in registros if r.get("tas_min") is not None and r["tas_min"] <= 30)
    con_eventos     = sum(1 for r in registros if r.get("eventos", 0) > 0)

    por_cereal = {}
    for r in registros:
        if r.get("estado_silo") != "Extraído":
            c = r.get("cereal", "Otro")
            por_cereal[c] = por_cereal.get(c, 0) + 1
    # ==========================================
    # RESUMEN COMERCIAL POR CEREAL
    # ==========================================
    conn2 = get_db()
    resumen_comercial = {}

    for cereal in por_cereal.keys():
        mercado = db_execute(conn2, """
            SELECT
                CASE WHEN usar_manual = 1 THEN pizarra_manual
                     ELSE pizarra_auto
                END AS pizarra,
                dolar
            FROM mercado
            WHERE cereal = ? AND empresa_id = ?
        """, (cereal, empresa_id)).fetchone()

        silos_cereal = [
            r for r in registros
            if r.get("cereal") == cereal
            and r.get("estado_silo") != "Extraído"
            and r.get("factor") is not None
        ]

        if not silos_cereal:
            continue

        kg_con_factor = []
        for r in silos_cereal:
            kg = r.get("kg_total") or 0
            if kg == 0:
                kg = (r.get("metros") or 0) * KG_POR_METRO.get(r.get("cereal", ""), KG_POR_METRO_DEFAULT)
            if kg > 0:
                kg_con_factor.append((kg, r["factor"]))

        if not kg_con_factor:
            continue

        kg_total_cereal = sum(k for k, _ in kg_con_factor)
        factor_pond = sum(k * f for k, f in kg_con_factor) / kg_total_cereal if kg_total_cereal > 0 else None

        precio_ars = None
        precio_usd = None
        if factor_pond and mercado and mercado["pizarra"] and mercado["dolar"]:
            precio_ars = round(mercado["pizarra"] * factor_pond, 2)
            precio_usd = round(precio_ars / mercado["dolar"], 2)

        resumen_comercial[cereal] = {
            "silos":       len(silos_cereal),
            "kg_total":    int(kg_total_cereal),
            "factor_pond": round(factor_pond * 100, 2) if factor_pond else None,
            "pizarra":     mercado["pizarra"] if mercado else None,
            "precio_ars":  precio_ars,
            "precio_usd":  precio_usd,
        }

    conn2.close()

    return render_template(
        "panel.html",
        registros=registros,
        puede_form=tiene_permiso("form"),
        puede_comercial=tiene_permiso("comercial"),
        puede_admin=tiene_permiso("admin"),
        resumen={
            "total_activos":       total_activos,
            "total_en_extraccion": total_en_extraccion,
            "total_extraidos":     total_extraidos,
            "con_alertas":     con_alertas,
            "con_eventos":     con_eventos,
            "por_cereal":      por_cereal,
        },
        resumen_comercial=resumen_comercial
    )


@panel_bp.route("/form")
@login_required
def form():

    if not tiene_permiso("form"):
        return acceso_denegado("form")

    return render_template(
        "form.html",
        puede_calado=tiene_permiso("calado")
    )


@panel_bp.route("/exportar_excel")
@login_required
def exportar_excel():

    if not tiene_permiso("panel"):
        return acceso_denegado("panel")

    conn = get_db()
    empresa_id = empresa_actual()

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    from openpyxl.utils import get_column_letter
    from io import BytesIO

    # =================================================================
    # NIVELES DE PERMISO
    # =================================================================
    es_admin     = tiene_permiso("admin") or getattr(current_user, 'es_superadmin', False) or getattr(current_user, 'rol', '') == 'admin_empresa'
    ve_comercial = es_admin or tiene_permiso("comercial")
    ve_calidad   = es_admin or tiene_permiso("calado") or tiene_permiso("laboratorio") or tiene_permiso("comercial")
    ve_form      = es_admin or tiene_permiso("form") or ve_calidad

    # =================================================================
    # PALETA DE COLORES
    # =================================================================
    C_VERDE_OSC  = "1B5E20"
    C_VERDE_MED  = "2E7D32"
    C_VERDE_CLR  = "E8F5E9"
    C_AZUL_OSC   = "0D47A1"
    C_AZUL_MED   = "1565C0"
    C_AZUL_CLR   = "E3F2FD"
    C_ROJO_CLR   = "FFEBEE"
    C_NARANJA    = "FFF3E0"
    C_GRIS       = "F5F5F5"
    C_GRIS2      = "E0E0E0"
    C_BLANCO     = "FFFFFF"
    C_MORADO     = "4A148C"
    C_MORADO_CLR = "F3E5F5"

    thin = Side(style="thin", color="BDBDBD")
    thick_bottom = Side(style="medium", color="757575")

    def _borde(bottom_thick=False):
        return Border(
            left=thin, right=thin, top=thin,
            bottom=thick_bottom if bottom_thick else thin
        )

    def _fill(c):
        return PatternFill("solid", fgColor=c)

    def _center():
        return Alignment(horizontal="center", vertical="center", wrap_text=True)

    def estilo_titulo(celda, color=C_VERDE_OSC):
        celda.font = Font(bold=True, color=C_BLANCO, size=14, name="Calibri")
        celda.fill = _fill(color)
        celda.alignment = _center()

    def estilo_subtitulo(celda, color=C_VERDE_MED):
        celda.font = Font(bold=True, color=C_BLANCO, size=11, name="Calibri")
        celda.fill = _fill(color)
        celda.alignment = _center()
        celda.border = _borde()

    def estilo_header(celda, color=C_VERDE_MED):
        celda.font = Font(bold=True, color=C_BLANCO, size=10, name="Calibri")
        celda.fill = _fill(color)
        celda.alignment = _center()
        celda.border = _borde(bottom_thick=True)

    def estilo_body(celda, bg=None, bold=False, num_fmt=None):
        if bg:
            celda.fill = _fill(bg)
        celda.border = _borde()
        celda.alignment = _center()
        celda.font = Font(name="Calibri", size=10, bold=bold)
        if num_fmt:
            celda.number_format = num_fmt

    def estilo_total(celda, num_fmt=None):
        celda.font = Font(bold=True, size=10, name="Calibri", color=C_VERDE_OSC)
        celda.fill = _fill(C_GRIS2)
        celda.border = _borde(bottom_thick=True)
        celda.alignment = _center()
        if num_fmt:
            celda.number_format = num_fmt

    def write_row(ws, row, vals, bg=None, bold=False, num_fmts=None):
        for c, v in enumerate(vals, 1):
            cell = ws.cell(row=row, column=c, value=v)
            nf = num_fmts[c - 1] if num_fmts and c - 1 < len(num_fmts) else None
            estilo_body(cell, bg=bg, bold=bold, num_fmt=nf)

    def write_headers(ws, row, headers, color=C_VERDE_MED):
        for c, h in enumerate(headers, 1):
            cell = ws.cell(row=row, column=c, value=h)
            estilo_header(cell, color)

    def set_col_widths(ws, widths):
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

    def _mapear_cereal(nombre):
        n = nombre.lower()
        if "soja" in n: return "Soja"
        if "maiz" in n or "maíz" in n: return "Maíz"
        if "trigo" in n: return "Trigo"
        if "girasol" in n: return "Girasol"
        if "sorgo" in n: return "Sorgo"
        return None

    # =================================================================
    # ██  EXCEL BÁSICO — solo panel sin permiso form  ██
    # =================================================================
    if not ve_form:

        silos_basico = db_execute(conn, """
            SELECT numero_qr, cereal, fecha_confeccion, estado_silo,
                   estado_grano, metros
            FROM silos
            WHERE empresa_id=?
            ORDER BY cereal, numero_qr
        """, (empresa_id,)).fetchall()

        # Agrupar por cereal
        por_cereal = {}
        for s in silos_basico:
            cer = s["cereal"] or "Otro"
            if cer not in por_cereal:
                por_cereal[cer] = []
            por_cereal[cer].append(s)

        wb = Workbook()
        wb.remove(wb.active)  # Eliminar hoja por defecto

        cereales_orden = ["Soja", "Maíz", "Trigo", "Girasol", "Sorgo"]

        # Primero los cereales en orden conocido
        for cereal in cereales_orden:
            if cereal not in por_cereal:
                continue
            silos_cer = por_cereal[cereal]
            ws = wb.create_sheet(cereal)
            ncols = 5

            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
            ws["A1"] = f"{cereal.upper()} — SILO BOLSAS"
            estilo_titulo(ws["A1"])
            ws.row_dimensions[1].height = 32

            ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)
            ws["A2"] = f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}  |  Silos: {len(silos_cer)}"
            ws["A2"].font = Font(italic=True, size=9, color="757575", name="Calibri")
            ws["A2"].alignment = _center()

            headers = ["QR", "Fecha Confección", "Estado Silo", "Estado Grano", "Metros"]
            write_headers(ws, 3, headers, C_VERDE_MED)

            row = 4
            for s in silos_cer:
                vals = [s["numero_qr"], s["fecha_confeccion"], s["estado_silo"], s["estado_grano"], s["metros"]]
                bg = C_GRIS if row % 2 == 0 else None
                write_row(ws, row, vals, bg=bg)
                row += 1

            set_col_widths(ws, [16, 20, 16, 16, 12])
            ws.auto_filter.ref = f"A3:E{row - 1}"
            ws.freeze_panes = "A4"

        # Cereales que no estén en la lista ordenada
        for cereal, silos_cer in por_cereal.items():
            if cereal in cereales_orden:
                continue
            ws = wb.create_sheet(cereal)
            ncols = 5
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
            ws["A1"] = f"{cereal.upper()} — SILO BOLSAS"
            estilo_titulo(ws["A1"])
            ws.row_dimensions[1].height = 32
            ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)
            ws["A2"] = f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}  |  Silos: {len(silos_cer)}"
            ws["A2"].font = Font(italic=True, size=9, color="757575", name="Calibri")
            ws["A2"].alignment = _center()
            headers = ["QR", "Fecha Confección", "Estado Silo", "Estado Grano", "Metros"]
            write_headers(ws, 3, headers, C_VERDE_MED)
            row = 4
            for s in silos_cer:
                vals = [s["numero_qr"], s["fecha_confeccion"], s["estado_silo"], s["estado_grano"], s["metros"]]
                bg = C_GRIS if row % 2 == 0 else None
                write_row(ws, row, vals, bg=bg)
                row += 1
            set_col_widths(ws, [16, 20, 16, 16, 12])
            ws.auto_filter.ref = f"A3:E{row - 1}"
            ws.freeze_panes = "A4"

        if not wb.sheetnames:
            ws = wb.create_sheet("Sin datos")
            ws["A1"] = "No hay silos registrados"

        registrar_auditoria(conn, current_user.id, empresa_id, "exportacion_excel", "Exportación básica (solo panel)", None)
        conn.close()

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        return send_file(
            output,
            as_attachment=True,
            download_name=f"silos_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    # =================================================================
    # ██  EXCEL COMPLETO — según nivel de permiso  ██
    # =================================================================

    silos = db_execute(conn, """
        SELECT s.*,
        (
            SELECT m.id
            FROM muestreos m
            WHERE m.numero_qr = s.numero_qr
            AND m.empresa_id = s.empresa_id
            ORDER BY m.id DESC
            LIMIT 1
        ) ultimo_muestreo
        FROM silos s
        WHERE s.empresa_id=?
        ORDER BY s.cereal, s.numero_qr
    """, (empresa_id,)).fetchall()

    mercado_rows = db_execute(conn, """
        SELECT cereal,
        CASE WHEN usar_manual=1 THEN pizarra_manual ELSE pizarra_auto END pizarra,
        dolar
        FROM mercado
        WHERE empresa_id=?
    """, (empresa_id,)).fetchall()

    mercado = {r["cereal"]: dict(r) for r in mercado_rows}

    try:
        matba_rows = db_execute(conn, """
            SELECT cereal, posicion, mes, precio, variacion
            FROM matba ORDER BY cereal, posicion
        """).fetchall()
    except Exception:
        try: conn.rollback()
        except: pass
        try:
            matba_rows = db_execute(conn, """
                SELECT cereal, posicion, mes, precio
                FROM matba ORDER BY cereal, posicion
            """).fetchall()
        except: matba_rows = []

    try:
        rofex_rows = db_execute(conn, """
            SELECT posicion, ajuste, variacion
            FROM rofex ORDER BY posicion
        """).fetchall()
    except Exception:
        try: conn.rollback()
        except: pass
        rofex_rows = []

    cereales = ["Soja", "Maíz", "Trigo", "Girasol", "Sorgo"]

    silo_data = []
    for s in silos:
        grado = None; factor = None; tas = None; fecha_est = None; fuente = "—"; kg = 0

        try:
            kg = db_execute(conn, """
                SELECT COALESCE(SUM(kg),0) total FROM llenado
                WHERE numero_qr=? AND empresa_id=?
            """, (s["numero_qr"], empresa_id)).fetchone()["total"]
        except Exception:
            try: conn.rollback()
            except: pass
            kg = 0

        if not kg:
            kg = (s["metros"] or 0) * KG_POR_METRO.get(s["cereal"], KG_POR_METRO_DEFAULT)

        if s["ultimo_muestreo"]:
            try:
                ana = db_execute(conn, """
                    SELECT grado, factor, tas FROM analisis
                    WHERE id_muestreo=? AND empresa_id=?
                """, (s["ultimo_muestreo"], empresa_id)).fetchall()
                grados_l = []; factores_l = []; tass_l = []
                for a in ana:
                    if a["grado"] is not None:
                        try: grados_l.append(int(a["grado"]))
                        except: pass
                    if a["factor"] is not None: factores_l.append(float(a["factor"]))
                    if a["tas"] is not None: tass_l.append(int(a["tas"]))
                if grados_l: grado = max(grados_l)
                if factores_l:
                    factor = round(sum(factores_l) / len(factores_l), 4); fuente = "Calado"
                if tass_l:
                    tas = min(tass_l)
                    f_row = db_execute(conn, """
                        SELECT fecha_muestreo FROM muestreos WHERE id=? AND empresa_id=?
                    """, (s["ultimo_muestreo"], empresa_id)).fetchone()
                    if f_row and f_row["fecha_muestreo"]:
                        try:
                            for fmt in ("%Y-%m-%d %H:%M", "%Y-%m-%d %H:%M:%S", "%Y-%m-%d"):
                                try: base = datetime.strptime(f_row["fecha_muestreo"], fmt); break
                                except ValueError: continue
                            fecha_est = (base + timedelta(days=tas)).strftime("%Y-%m-%d")
                        except: pass
            except Exception:
                try: conn.rollback()
                except: pass

        if factor is None:
            try:
                cargas = db_execute(conn, """
                    SELECT kg, factor, tas, fecha FROM llenado
                    WHERE numero_qr=? AND empresa_id=? ORDER BY fecha DESC
                """, (s["numero_qr"], empresa_id)).fetchall()
                if cargas:
                    cargas_cf = [c for c in cargas if c["factor"] is not None]
                    if cargas_cf:
                        kg_pond = sum(float(c["kg"] or 0) for c in cargas_cf)
                        if kg_pond > 0:
                            factor = round(sum(float(c["factor"]) * float(c["kg"] or 0) for c in cargas_cf) / kg_pond, 4)
                        else:
                            fl = [float(c["factor"]) for c in cargas_cf]
                            factor = round(sum(fl) / len(fl), 4)
                        fuente = "Llenado"
                    tas_vals = [int(c["tas"]) for c in cargas if c["tas"] is not None]
                    if tas_vals:
                        tas = min(tas_vals)
                        if cargas[0]["fecha"]:
                            try:
                                for fmt in ("%Y-%m-%d %H:%M", "%Y-%m-%d %H:%M:%S", "%Y-%m-%d"):
                                    try: base = datetime.strptime(cargas[0]["fecha"], fmt); break
                                    except ValueError: continue
                                fecha_est = (base + timedelta(days=tas)).strftime("%Y-%m-%d")
                            except: pass
            except Exception:
                try: conn.rollback()
                except: pass

        cereal = s["cereal"]; merc = mercado.get(cereal)
        precio_ars = None; precio_usd = None
        if ve_comercial and factor and merc and merc["pizarra"]:
            precio_ars = round(merc["pizarra"] * factor, 2)
            if merc["dolar"]: precio_usd = round(precio_ars / merc["dolar"], 2)

        # ── datos de vaciado ──
        kg_vaciado = 0; factor_vaciado = None; dif_kg = None; dif_factor = None
        camionadas_silo = []
        if s["estado_silo"] in ("Extraído", "En extracción"):
            try:
                cam_rows = conn.execute(
                    "SELECT * FROM vaciado WHERE numero_qr=? AND empresa_id=? ORDER BY nro_camion ASC",
                    (s["numero_qr"], empresa_id)
                ).fetchall()
                camionadas_silo = [dict(c) for c in cam_rows]
                kg_vaciado = int(sum(float(c["kg"] or 0) for c in camionadas_silo if c.get("kg")))
                kg_sum_vac = sum(float(c["kg"] or 0) for c in camionadas_silo if c.get("factor") and c.get("kg"))
                if kg_sum_vac > 0:
                    factor_vaciado = round(
                        sum(float(c["factor"]) * float(c["kg"] or 0) for c in camionadas_silo if c.get("factor") and c.get("kg"))
                        / kg_sum_vac, 4
                    )
            except Exception:
                try: conn.rollback()
                except: pass
        if s["estado_silo"] == "Extraído" and kg > 0:
            dif_kg = kg_vaciado - int(kg)
        if factor and factor_vaciado:
            dif_factor = round((factor_vaciado - factor) * 100, 3)

        silo_data.append({
            "numero_qr": s["numero_qr"], "cereal": cereal,
            "fecha_confeccion": s["fecha_confeccion"], "estado_silo": s["estado_silo"],
            "fecha_extraccion": s["fecha_extraccion"] if "fecha_extraccion" in s.keys() else None,
            "estado_grano": s["estado_grano"], "metros": s["metros"],
            "kg": int(kg), "grado": grado, "factor": factor, "tas": tas,
            "fecha_est": fecha_est, "fuente": fuente,
            "precio_ars": precio_ars, "precio_usd": precio_usd,
            "kg_vaciado": kg_vaciado, "factor_vaciado": factor_vaciado,
            "dif_kg": dif_kg, "dif_factor": dif_factor,
            "camionadas": camionadas_silo,
        })

    wb = Workbook(); wb.remove(wb.active)

    # HOJA RESUMEN
    ws = wb.create_sheet("Resumen", 0)
    ncols = 13  # se recalcula después de saber ve_comercial
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    ws["A1"] = "INFORME SILO BOLSAS"; estilo_titulo(ws["A1"]); ws.row_dimensions[1].height = 32
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)
    ws["A2"] = f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    ws["A2"].font = Font(italic=True, size=9, color="757575", name="Calibri"); ws["A2"].alignment = _center()

    row = 4
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    ws.cell(row=row, column=1, value="RESUMEN POR CEREAL"); estilo_subtitulo(ws.cell(row=row, column=1), C_VERDE_MED); row += 1
    # Columnas del resumen según permiso
    res_headers = ["Cereal", "Activos", "En Extracción", "Extraídos", "KG Llenado", "KG Vaciado", "Dif. KG", "Factor Pond. %", "Silos c/Factor"]
    res_fmts    = [None, "#,##0", "#,##0", "#,##0", "#,##0", "#,##0", "+#,##0;-#,##0", "0.00", None]
    if ve_comercial:
        res_headers += ["Pizarra ARS/TN", "Pizarra USD/TN", "Valor Stock ARS", "Valor Stock USD"]
        res_fmts    += ["#,##0.00", "#,##0.00", "#,##0", "#,##0"]
    ncols = len(res_headers)
    write_headers(ws, row, res_headers, C_VERDE_MED); row += 1

    tot_activos = tot_en_extr = tot_extraidos = tot_kg = tot_kg_vac = tot_ars = tot_usd = tot_con_factor = tot_silos_activos = 0
    for cereal in cereales:
        activos    = [d for d in silo_data if d["cereal"] == cereal and d["estado_silo"] not in ("Extraído", "En extracción")]
        en_extr    = [d for d in silo_data if d["cereal"] == cereal and d["estado_silo"] == "En extracción"]
        extraidos  = [d for d in silo_data if d["cereal"] == cereal and d["estado_silo"] == "Extraído"]
        if not activos and not en_extr and not extraidos: continue
        kg_total = sum(d["kg"] for d in activos + en_extr)
        kg_vac_total = sum(d["kg_vaciado"] for d in extraidos)
        con_factor = [d for d in activos + en_extr if d["factor"] is not None]; n_con_factor = len(con_factor)
        factor_pond = None
        if con_factor:
            kg_f = sum(d["kg"] for d in con_factor)
            factor_pond = sum(d["factor"] * d["kg"] for d in con_factor) / kg_f if kg_f > 0 else sum(d["factor"] for d in con_factor) / len(con_factor)
        merc = mercado.get(cereal)
        piz_ars = merc["pizarra"] if merc else None; dolar = merc["dolar"] if merc else None
        piz_usd = round(piz_ars / dolar, 2) if piz_ars and dolar else None
        val_ars = val_usd = None
        if factor_pond and piz_ars and kg_total > 0:
            val_ars = round(piz_ars * factor_pond * kg_total / 1000, 2)
            if dolar: val_usd = round(val_ars / dolar, 2)
        dif_kg_cer = (kg_vac_total - sum(d["kg"] for d in extraidos)) if extraidos else None
        vals = [cereal, len(activos), len(en_extr), len(extraidos), kg_total,
                kg_vac_total if extraidos else None, dif_kg_cer,
                round(factor_pond * 100, 2) if factor_pond else None,
                f"{n_con_factor}/{len(activos + en_extr)}"]
        if ve_comercial:
            vals += [piz_ars, piz_usd, val_ars, val_usd]
        bg = C_GRIS if row % 2 == 0 else None
        write_row(ws, row, vals, bg=bg, num_fmts=res_fmts)
        # color dif_kg en la celda col 7
        cell_dif = ws.cell(row=row, column=7)
        if dif_kg_cer is not None:
            cell_dif.font = Font(name="Calibri", size=10, bold=True,
                color=C_VERDE_MED if dif_kg_cer >= 0 else "C62828")
        row += 1
        tot_activos += len(activos); tot_en_extr += len(en_extr); tot_extraidos += len(extraidos)
        tot_kg += kg_total; tot_kg_vac += kg_vac_total
        tot_ars += (val_ars or 0); tot_usd += (val_usd or 0)
        tot_con_factor += n_con_factor; tot_silos_activos += len(activos + en_extr)

    tot_dif = (tot_kg_vac - sum(d["kg"] for d in silo_data if d["estado_silo"] == "Extraído")) if tot_extraidos else None
    total_vals = ["TOTAL", tot_activos, tot_en_extr, tot_extraidos, tot_kg,
                  tot_kg_vac if tot_extraidos else None, tot_dif,
                  None, f"{tot_con_factor}/{tot_silos_activos}"]
    if ve_comercial:
        total_vals += [None, None, tot_ars if tot_ars else None, tot_usd if tot_usd else None]
    for c, v in enumerate(total_vals, 1):
        cell = ws.cell(row=row, column=c, value=v)
        estilo_total(cell, num_fmt=res_fmts[c-1] if c-1 < len(res_fmts) else None)
        if c == 7 and tot_dif is not None:
            cell.font = Font(bold=True, size=10, name="Calibri",
                color=C_VERDE_MED if tot_dif >= 0 else "C62828")
    row += 2

    # Construir diccionarios matba_por_cereal y rofex_dict
    # (disponibles para análisis interactivo Y hoja Mejor Combo)
    matba_por_cereal = {}
    for m in (matba_rows or []):
        cer_simple = _mapear_cereal(m["cereal"])
        if not cer_simple: continue
        if cer_simple not in matba_por_cereal:
            matba_por_cereal[cer_simple] = []
        try: contrato = m["mes"]
        except: contrato = None
        matba_por_cereal[cer_simple].append({
            "posicion": m["posicion"],
            "mes": contrato,
            "precio": m["precio"],
        })

    rofex_dict = {r["posicion"]: r["ajuste"] for r in rofex_rows} if rofex_rows else {}

    # ── Sección: Análisis de Precios con Futuros (solo ve_comercial) ───────────
    if ve_comercial:
        from openpyxl.worksheet.datavalidation import DataValidation

        # Hoja de datos auxiliar para los dropdowns (oculta)
        ws_aux = wb.create_sheet("_AuxDropdowns")
        ws_aux.sheet_state = "hidden"

        # Escribir opciones ROFEX en columna A de la hoja aux
        rofex_opts = list(rofex_dict.keys())
        ws_aux["A1"] = "ROFEX"
        for i, pos in enumerate(rofex_opts, 2):
            ws_aux.cell(row=i, column=1, value=pos)
        rofex_range = f"_AuxDropdowns!$A$2:$A${len(rofex_opts)+1}" if rofex_opts else None

        # Escribir opciones MATBA por cereal en columnas B+ de la hoja aux
        matba_col_map = {}  # cereal -> letra de columna en aux
        aux_col = 2
        for cer in cereales:
            opts = matba_por_cereal.get(cer, [])
            if not opts: continue
            ws_aux.cell(row=1, column=aux_col, value=f"MATBA_{cer}")
            for i, m in enumerate(opts, 2):
                label = f"{m['posicion']} ({m['mes']})" if m.get('mes') else m['posicion']
                ws_aux.cell(row=i, column=aux_col, value=label)
            col_letter = get_column_letter(aux_col)
            matba_col_map[cer] = (col_letter, len(opts), opts)
            aux_col += 1

        # ── Tabla de precios con selección interactiva ───────────────────────────
        row += 1
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
        ws.cell(row=row, column=1, value="ANÁLISIS DE PRECIOS — FUTUROS vs PIZARRA")
        estilo_subtitulo(ws.cell(row=row, column=1), C_AZUL_OSC); row += 1

        # Fila de selección global de dólar ROFEX
        if rofex_opts:
            ws.cell(row=row, column=1, value="Dólar ROFEX a usar:")
            ws.cell(row=row, column=1).font = Font(bold=True, size=10, name="Calibri", color=C_MORADO)
            sel_rofex_cell = ws.cell(row=row, column=2, value=rofex_opts[0])
            sel_rofex_cell.fill = _fill("EDE7F6")
            sel_rofex_cell.font = Font(bold=True, size=10, name="Calibri", color=C_MORADO)
            sel_rofex_cell.border = Border(left=Side(style="medium", color=C_MORADO),
                                            right=Side(style="medium", color=C_MORADO),
                                            top=Side(style="medium", color=C_MORADO),
                                            bottom=Side(style="medium", color=C_MORADO))
            sel_rofex_cell.alignment = _center()
            sel_rofex_ref = f"B{row}"
            dv_rofex = DataValidation(type="list", formula1=f"_AuxDropdowns!$A$2:$A${len(rofex_opts)+1}", allow_blank=False, showDropDown=False)
            dv_rofex.sqref = sel_rofex_ref
            ws.add_data_validation(dv_rofex)
            ws.cell(row=row, column=3, value="← Seleccioná la posición de dólar futuro").font = Font(italic=True, size=9, color="9E9E9E", name="Calibri")
            row += 1

        # Guardar referencias de fila para cada cereal (para la fórmula del mejor precio)
        sel_matba_refs = {}  # cereal -> cell ref de la celda de selección MATBA

        # Encabezados de la tabla
        row += 1
        anl_headers = ["Cereal", "KG Stock", "Pizarra ARS", "Dólar Hoy", "Precio Pizarra USD/TN",
                       "Contrato MATBA", "Precio Futuro USD/TN", "Dif. vs Pizarra %",
                       "Dólar ROFEX sel.", "Precio Futuro ARS/TN", "Valor Stock Futuro ARS", "Mejor Precio"]
        anl_fmts = [None, "#,##0", "#,##0.00", "#,##0.00", "#,##0.00",
                    None, "#,##0.00", "+0.00%;-0.00%",
                    "#,##0.00", "#,##0.00", "#,##0", None]
        write_headers(ws, row, anl_headers, C_AZUL_MED); row += 1
        header_row = row - 1

        for cer in cereales:
            merc = mercado.get(cer)
            opts_info = matba_por_cereal.get(cer, [])
            if not merc and not opts_info: continue

            piz_ars  = merc["pizarra"] if merc else None
            dol_hoy  = merc["dolar"]   if merc else None
            piz_usd  = round(piz_ars / dol_hoy, 2) if piz_ars and dol_hoy else None
            kg_stock = sum(d["kg"] for d in silo_data if d["cereal"] == cer and d["estado_silo"] not in ("Extraído",))

            # celda de selección del contrato MATBA para este cereal
            col_info = matba_col_map.get(cer)
            matba_sel_cell_ref = None
            if col_info:
                col_letter, n_opts, opts = col_info
                sel_cell = ws.cell(row=row, column=6)
                first_opt = f"{opts[0]['posicion']} ({opts[0]['mes']})" if opts[0].get('mes') else opts[0]['posicion']
                sel_cell.value = first_opt
                sel_cell.fill = _fill(C_AZUL_CLR)
                sel_cell.font = Font(bold=True, size=10, name="Calibri", color=C_AZUL_OSC)
                sel_cell.border = Border(left=Side(style="medium", color=C_AZUL_MED),
                                          right=Side(style="medium", color=C_AZUL_MED),
                                          top=Side(style="medium", color=C_AZUL_MED),
                                          bottom=Side(style="medium", color=C_AZUL_MED))
                sel_cell.alignment = _center()
                matba_sel_cell_ref = f"F{row}"
                dv_matba = DataValidation(
                    type="list",
                    formula1=f"_AuxDropdowns!${col_letter}$2:${col_letter}${n_opts+1}",
                    allow_blank=True, showDropDown=False
                )
                dv_matba.sqref = matba_sel_cell_ref
                ws.add_data_validation(dv_matba)
                sel_matba_refs[cer] = matba_sel_cell_ref

                # Precio futuro: BUSCARV del contrato seleccionado en la hoja aux
                # Construir tabla de lookup en aux: col_letter tiene labels, col_letter+1 tiene precios
                price_col = aux_col  # próxima columna disponible
                ws_aux.cell(row=1, column=price_col, value=f"PRECIO_{cer}")
                for i, m in enumerate(opts, 2):
                    label = f"{m['posicion']} ({m['mes']})" if m.get('mes') else m['posicion']
                    ws_aux.cell(row=i, column=price_col, value=m["precio"])
                price_col_letter = get_column_letter(price_col)
                aux_col += 1

                # Fórmula INDICE+COINCIDIR para buscar el precio del contrato seleccionado
                lookup_range   = f"_AuxDropdowns!${col_letter}$2:${col_letter}${n_opts+1}"
                price_range    = f"_AuxDropdowns!${price_col_letter}$2:${price_col_letter}${n_opts+1}"
                fut_usd_formula = f"=IFERROR(INDEX({price_range},MATCH(F{row},{lookup_range},0)),"")"

                ws.cell(row=row, column=7).value = fut_usd_formula
                estilo_body(ws.cell(row=row, column=7), num_fmt="#,##0.00")

                # Diferencia % vs pizarra
                if piz_usd:
                    ws.cell(row=row, column=8).value = f'=IFERROR((G{row}-{piz_usd})/{piz_usd},"")'
                    estilo_body(ws.cell(row=row, column=8), num_fmt="+0.00%;-0.00%")
                    # Formato condicional de color para dif
                    dif_cell = ws.cell(row=row, column=8)
                    dif_cell.fill = _fill(C_GRIS)  # placeholder; color real depende del valor calculado
            else:
                for col in (6, 7, 8): estilo_body(ws.cell(row=row, column=col))
                ws.cell(row=row, column=6).value = "Sin datos MATBA"

            # Dólar ROFEX seleccionado: BUSCARV en _AuxDropdowns col A → ajuste
            rofex_price_col = aux_col
            ws_aux.cell(row=1, column=rofex_price_col, value="ROFEX_AJUSTE")
            for i, (pos, ajuste) in enumerate(rofex_dict.items(), 2):
                ws_aux.cell(row=i, column=rofex_price_col, value=ajuste)
            rofex_price_col_letter = get_column_letter(rofex_price_col)

            if sel_rofex_ref and rofex_opts:
                rofex_lookup = f"_AuxDropdowns!$A$2:$A${len(rofex_opts)+1}"
                rofex_price_lookup = f"_AuxDropdowns!${rofex_price_col_letter}$2:${rofex_price_col_letter}${len(rofex_opts)+1}"
                dol_rofex_formula = f"=IFERROR(INDEX({rofex_price_lookup},MATCH({sel_rofex_ref},{rofex_lookup},0)),{dol_hoy or 'NA()'})"
                ws.cell(row=row, column=9).value = dol_rofex_formula
                estilo_body(ws.cell(row=row, column=9), num_fmt="#,##0.00")
                # Precio futuro en ARS = precio futuro USD * dolar ROFEX seleccionado
                ws.cell(row=row, column=10).value = f"=IFERROR(G{row}*I{row},"")"
                estilo_body(ws.cell(row=row, column=10), num_fmt="#,##0.00")
                # Valor stock con precio futuro ARS
                ws.cell(row=row, column=11).value = f"=IFERROR(J{row}*{kg_stock}/1000,"")"
                estilo_body(ws.cell(row=row, column=11), num_fmt="#,##0")
            else:
                for col in (9, 10, 11): estilo_body(ws.cell(row=row, column=col))

            # Mejor precio: comparar pizarra USD vs futuro USD
            if piz_usd:
                ws.cell(row=row, column=12).value = (
                    f'=IFERROR(IF(G{row}="","Pizarra",'
                    f'IF(G{row}>{piz_usd},"▲ Futuro (" & TEXT(G{row},"0.00") & " USD)","▼ Pizarra (" & TEXT({piz_usd},"0.00") & " USD)")),"Pizarra")'
                )
                mejor_cell = ws.cell(row=row, column=12)
                estilo_body(mejor_cell)
                mejor_cell.font = Font(bold=True, size=10, name="Calibri")
            else:
                estilo_body(ws.cell(row=row, column=12))

            # Columnas fijas
            bg = C_GRIS if row % 2 == 0 else None
            fixed_vals = [cer, kg_stock, piz_ars, dol_hoy, piz_usd]
            fixed_fmts = [None, "#,##0", "#,##0.00", "#,##0.00", "#,##0.00"]
            for c_idx, (val, fmt) in enumerate(zip(fixed_vals, fixed_fmts), 1):
                cell = ws.cell(row=row, column=c_idx, value=val)
                estilo_body(cell, bg=bg, num_fmt=fmt)
            row += 1

        # Nota informativa
        row += 1
        ws.cell(row=row, column=1, value="💡 Seleccioná el contrato MATBA en columna F y el dólar ROFEX en la celda de selección para ver el análisis actualizado.").font = Font(italic=True, size=9, color="616161", name="Calibri")
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)

        set_col_widths(ws, [14, 10, 13, 11, 14, 14, 14, 13, 14, 14, 16, 16, 13])
        widths_res = [14, 10, 13, 11, 14, 14, 14, 13, 13]
        if ve_comercial: widths_res += [14, 14, 16, 16]
        set_col_widths(ws, widths_res)
        col_last_res = get_column_letter(len(widths_res))
        ws.auto_filter.ref = f"A5:{col_last_res}{row - 2}"; ws.freeze_panes = "A6"

    # HOJAS POR CEREAL (admin)
    for cereal in cereales:
        data_cereal = [d for d in silo_data if d["cereal"] == cereal]
        if not data_cereal: continue
        ws = wb.create_sheet(cereal); ncols_c = 13
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols_c)
        ws["A1"] = f"{cereal.upper()} — DETALLE SILOS"; estilo_titulo(ws["A1"]); ws.row_dimensions[1].height = 30
        activos_c = [d for d in data_cereal if d["estado_silo"] != "Extraído"]
        con_factor_c = [d for d in activos_c if d["factor"] is not None]
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols_c)
        ws["A2"] = f"Activos: {len(activos_c)}  |  Extraídos: {len(data_cereal) - len(activos_c)}  |  Con factor: {len(con_factor_c)}/{len(activos_c)}"
        ws["A2"].font = Font(italic=True, size=9, color="616161", name="Calibri"); ws["A2"].alignment = _center()
        # Columnas según nivel de permiso
        # Base (ve_form): QR, Fecha Conf, Estado, Estado Grano, Metros, KG Ll, Fecha Cierre, KG Vaciado
        # + ve_calidad:   Grado, Factor Ll%, TAS, Fecha Est, Factor Vac%, Dif KG, Dif Factor
        # + ve_comercial: Precio ARS, Precio USD, Fuente
        headers_base    = ["QR", "Fecha Conf.", "Estado", "Estado Grano", "Metros", "KG Llenado", "Fecha Cierre", "KG Vaciado"]
        fmts_base       = [None, None, None, None, "0.0", "#,##0", None, "#,##0"]
        headers_calidad = ["Grado", "Factor Ll. %", "TAS", "Fecha Est.", "Factor Vac. %", "Dif. KG", "Dif. Factor %"]
        fmts_calidad    = [None, "0.00", "#,##0", None, "0.00", "+#,##0;-#,##0", "+0.000;-0.000"]
        headers_precio  = ["Precio ARS/TN", "Precio USD/TN", "Fuente"]
        fmts_precio     = ["#,##0.00", "#,##0.00", None]

        headers = headers_base
        hdr_fmts = fmts_base
        if ve_calidad:
            headers  += headers_calidad
            hdr_fmts += fmts_calidad
        if ve_comercial:
            headers  += headers_precio
            hdr_fmts += fmts_precio

        ncols_c = len(headers)
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols_c)
        write_headers(ws, 3, headers, C_VERDE_MED)

        row = 4; sum_kg_ll = 0; sum_kg_vac = 0; sum_factores_ll = []; sum_factores_vac = []
        for d in data_cereal:
            fac_ll_pct  = round(d["factor"] * 100, 2) if d["factor"] else None
            fac_vac_pct = round(d["factor_vaciado"] * 100, 2) if d.get("factor_vaciado") else None

            vals = [d["numero_qr"], d["fecha_confeccion"], d["estado_silo"], d["estado_grano"], d["metros"],
                    d["kg"], d.get("fecha_extraccion"), d.get("kg_vaciado") or None]
            if ve_calidad:
                vals += [d["grado"], fac_ll_pct, d["tas"], d["fecha_est"], fac_vac_pct,
                         d.get("dif_kg"), d.get("dif_factor")]
            if ve_comercial:
                vals += [d["precio_ars"], d["precio_usd"], d["fuente"]]

            if d["estado_silo"] == "Extraído":                           bg = C_ROJO_CLR
            elif d["estado_silo"] == "En extracción":                    bg = "FFF8E1"
            elif d["tas"] is not None and d["tas"] <= 30 and ve_calidad: bg = C_NARANJA
            elif d["tas"] is not None and d["tas"] > 60  and ve_calidad: bg = C_VERDE_CLR
            else:                                                         bg = C_GRIS if row % 2 == 0 else None
            write_row(ws, row, vals, bg=bg, num_fmts=hdr_fmts)

            # color dif_kg y dif_factor (solo si ve_calidad)
            if ve_calidad:
                col_dif_kg  = 14
                col_dif_fac = 15
                for col_idx, val in [(col_dif_kg, d.get("dif_kg")), (col_dif_fac, d.get("dif_factor"))]:
                    if val is not None:
                        ws.cell(row=row, column=col_idx).font = Font(
                            name="Calibri", size=10, bold=True,
                            color=C_VERDE_MED if val >= 0 else "C62828")

            if d["estado_silo"] != "Extraído":
                sum_kg_ll += d["kg"]
                if d["factor"] and ve_calidad: sum_factores_ll.append(d["factor"])
            if d["estado_silo"] == "Extraído":
                sum_kg_vac += (d.get("kg_vaciado") or 0)
                if d.get("factor_vaciado") and ve_calidad: sum_factores_vac.append(d["factor_vaciado"])
            row += 1

        avg_factor_ll  = round(sum(sum_factores_ll)  / len(sum_factores_ll)  * 100, 2) if sum_factores_ll  else None
        avg_factor_vac = round(sum(sum_factores_vac) / len(sum_factores_vac) * 100, 2) if sum_factores_vac else None
        tot_vals = ["TOTALES", None, None, None, None, sum_kg_ll, None, sum_kg_vac if sum_kg_vac else None]
        if ve_calidad: tot_vals += [None, avg_factor_ll, None, None, avg_factor_vac, None, None]
        if ve_comercial: tot_vals += [None, None, None]
        for c, v in enumerate(tot_vals, 1):
            cell = ws.cell(row=row, column=c, value=v)
            estilo_total(cell, num_fmt=hdr_fmts[c-1] if c-1 < len(hdr_fmts) else None)

        # anchos adaptados al nivel
        widths_base    = [14, 16, 14, 14, 9, 13, 16, 13]
        widths_calidad = [9, 12, 9, 16, 13, 13, 13]
        widths_precio  = [14, 14, 10]
        widths = widths_base + (widths_calidad if ve_calidad else []) + (widths_precio if ve_comercial else [])
        set_col_widths(ws, widths)
        col_last = get_column_letter(ncols_c)
        ws.auto_filter.ref = f"A3:{col_last}{row}"; ws.freeze_panes = "A4"

    # HOJA VACIADO — detalle de todas las camionadas
    todas_camionadas = []
    for d in silo_data:
        for c in d.get("camionadas", []):
            todas_camionadas.append({**c, "cereal": d["cereal"]})

    if todas_camionadas:
        ws_vac = wb.create_sheet("Vaciado"); ncols_v = 15
        ws_vac.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols_v)
        ws_vac["A1"] = "DETALLE DE VACIADO — CAMIONADAS"; estilo_titulo(ws_vac["A1"], C_AZUL_OSC)
        ws_vac.row_dimensions[1].height = 30
        ws_vac.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols_v)
        ws_vac["A2"] = f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}  |  Total camionadas: {len(todas_camionadas)}"
        ws_vac["A2"].font = Font(italic=True, size=9, color="757575", name="Calibri"); ws_vac["A2"].alignment = _center()

        vac_headers = ["QR Silo", "Cereal", "Nº Camión", "Patente", "Fecha",
                       "KG", "Humedad %", "Factor %", "TAS", "Temperatura",
                       "Insectos", "Destino", "Obs.", "Estado", "Fecha Cierre Silo"]
        vac_fmts    = [None, None, "#,##0", None, None,
                       "#,##0", "0.0", "0.00", "#,##0", "0.0",
                       None, None, None, None, None]
        write_headers(ws_vac, 3, vac_headers, C_AZUL_MED)

        row = 4
        for c in todas_camionadas:
            completado = c.get("completado", 0)
            destino    = (c.get("destino") or "").upper()
            bg = C_ROJO_CLR if not completado else (C_AZUL_CLR if destino == "PUERTO" else C_MORADO_CLR if destino == "PLANTA" else (C_GRIS if row % 2 == 0 else None))
            estado_cam = "Completa" if completado else "Pendiente lab."

            # buscar fecha_extraccion del silo padre
            silo_padre = next((d for d in silo_data if d["numero_qr"] == c.get("numero_qr")), None)
            fecha_cierre = silo_padre.get("fecha_extraccion") if silo_padre else None

            vals = [
                c.get("numero_qr"), c.get("cereal"), c.get("nro_camion"), c.get("patente"),
                c.get("fecha"),
                c.get("kg"), c.get("humedad"),
                round(float(c["factor"]) * 100, 2) if c.get("factor") else None,
                c.get("tas"), c.get("temperatura"),
                "Sí" if c.get("insectos") else "No",
                destino or "—", c.get("obs") or "—", estado_cam, fecha_cierre,
            ]
            write_row(ws_vac, row, vals, bg=bg, num_fmts=vac_fmts)
            # negrita en destino
            ws_vac.cell(row=row, column=12).font = Font(name="Calibri", size=10, bold=True)
            row += 1

        # fila totales
        tot_kg_cam = sum(float(c.get("kg") or 0) for c in todas_camionadas if c.get("kg"))
        for c_idx in range(1, ncols_v + 1):
            cell = ws_vac.cell(row=row, column=c_idx)
            if c_idx == 1: cell.value = "TOTAL"
            elif c_idx == 6: cell.value = int(tot_kg_cam)
            estilo_total(cell, num_fmt=vac_fmts[c_idx - 1] if c_idx - 1 < len(vac_fmts) else None)

        set_col_widths(ws_vac, [14, 10, 10, 12, 18, 12, 11, 10, 9, 12, 10, 10, 20, 14, 18])
        ws_vac.auto_filter.ref = f"A3:O{row}"
        ws_vac.freeze_panes = "A4"

    # HOJA MEJOR COMBO MATBA+ROFEX (solo ve_comercial)
    if ve_comercial and matba_rows and rofex_rows:
        ws_combo = wb.create_sheet("Mejor Combo")
        ncols_combo = 8
        ws_combo.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols_combo)
        ws_combo["A1"] = "MEJOR COMBINACIÓN MATBA + ROFEX POR CEREAL"
        estilo_titulo(ws_combo["A1"], C_AZUL_OSC); ws_combo.row_dimensions[1].height = 30

        ws_combo.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols_combo)
        ws_combo["A2"] = (
            f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}  |  "
            "Muestra todas las combinaciones posibles y resalta la que da mayor precio ARS/TN"
        )
        ws_combo["A2"].font = Font(italic=True, size=9, color="757575", name="Calibri")
        ws_combo["A2"].alignment = _center()

        combo_hdrs = ["Cereal", "Contrato MATBA", "Precio Futuro USD/TN",
                      "Posición ROFEX", "Dólar Futuro ARS", "Precio Final ARS/TN",
                      "vs Pizarra Hoy", "¿Mejor opción?"]
        combo_fmts = [None, None, "#,##0.00", None, "#,##0.00", "#,##0.00", "+0.00%;-0.00%", None]
        write_headers(ws_combo, 3, combo_hdrs, C_AZUL_OSC)

        row = 4
        for cer in cereales:
            merc = mercado.get(cer)
            if not merc or not merc.get("pizarra") or not merc.get("dolar"):
                continue
            piz_ars = merc["pizarra"]; dol_hoy = merc["dolar"]
            piz_usd = piz_ars / dol_hoy
            piz_ars_final = piz_ars  # precio actual = pizarra (sin futuro)
            opts = matba_por_cereal.get(cer, [])
            if not opts:
                continue

            # Calcular todas las combinaciones
            combos = []
            for m in opts:
                if not m.get("precio"): continue
                for pos, ajuste in rofex_dict.items():
                    if not ajuste: continue
                    precio_ars_combo = round(m["precio"] * ajuste, 2)
                    vs_piz = (precio_ars_combo - piz_ars_final) / piz_ars_final if piz_ars_final else 0
                    label_m = f"{m['posicion']} ({m['mes']})" if m.get('mes') else m['posicion']
                    combos.append({
                        "label_m": label_m, "precio_usd": m["precio"],
                        "pos_rofex": pos, "ajuste": ajuste,
                        "precio_ars": precio_ars_combo, "vs_piz": vs_piz
                    })

            if not combos: continue
            mejor = max(combos, key=lambda x: x["precio_ars"])

            # Fila de pizarra actual como referencia
            ws_combo.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols_combo)
            ref_cell = ws_combo.cell(row=row, column=1,
                value=f"── {cer.upper()} — Pizarra hoy: {piz_ars:,.2f} ARS/TN  ({piz_usd:.2f} USD/TN) ──")
            ref_cell.fill = _fill("E3F2FD")
            ref_cell.font = Font(bold=True, size=10, name="Calibri", color=C_AZUL_OSC)
            ref_cell.alignment = _center()
            row += 1

            # Todas las combinaciones, mejor resaltada
            for c in sorted(combos, key=lambda x: x["precio_ars"], reverse=True):
                es_mejor = c is mejor
                bg = "FFF9C4" if es_mejor else (C_GRIS if row % 2 == 0 else None)
                vals = [cer, c["label_m"], c["precio_usd"], c["pos_rofex"],
                        c["ajuste"], c["precio_ars"], c["vs_piz"],
                        "⭐ MEJOR COMBO" if es_mejor else ""]
                write_row(ws_combo, row, vals, bg=bg, num_fmts=combo_fmts)
                # bold + color en precio final y vs pizarra
                cell_precio = ws_combo.cell(row=row, column=6)
                cell_vs     = ws_combo.cell(row=row, column=7)
                if es_mejor:
                    cell_precio.font = Font(bold=True, size=11, name="Calibri", color="1B5E20")
                    cell_vs.font     = Font(bold=True, size=10, name="Calibri",
                        color=C_VERDE_MED if c["vs_piz"] >= 0 else "C62828")
                    ws_combo.cell(row=row, column=8).font = Font(bold=True, size=10,
                        name="Calibri", color="1B5E20")
                else:
                    color_vs = C_VERDE_MED if c["vs_piz"] >= 0 else "C62828"
                    cell_vs.font = Font(size=10, name="Calibri", color=color_vs)
                row += 1
            row += 1  # separador entre cereales

        set_col_widths(ws_combo, [14, 22, 18, 16, 16, 18, 15, 16])
        ws_combo.auto_filter.ref = f"A3:H{row - 1}"
        ws_combo.freeze_panes = "A4"

    # HOJA ROFEX
    if rofex_rows:
        ws = wb.create_sheet("ROFEX"); ncols_r = 3
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols_r)
        ws["A1"] = "DÓLAR FUTURO — ROFEX"; estilo_titulo(ws["A1"], C_MORADO); ws.row_dimensions[1].height = 30
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols_r)
        ws["A2"] = f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
        ws["A2"].font = Font(italic=True, size=9, color="757575", name="Calibri"); ws["A2"].alignment = _center()
        write_headers(ws, 3, ["Posición", "Ajuste", "Variación"], C_MORADO)
        row = 4
        for r in rofex_rows:
            variacion = r["variacion"] if r["variacion"] is not None else 0
            bg = C_VERDE_CLR if variacion > 0 else (C_ROJO_CLR if variacion < 0 else (C_GRIS if row % 2 == 0 else None))
            write_row(ws, row, [r["posicion"], r["ajuste"], variacion], bg=bg, num_fmts=[None, "#,##0.00", "0.00"]); row += 1
        set_col_widths(ws, [18, 16, 14]); ws.auto_filter.ref = f"A3:C{row}"; ws.freeze_panes = "A4"

    # HOJA MATBA
    if matba_rows:
        ws = wb.create_sheet("Futuros MATBA"); ncols_f = 9
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols_f)
        ws["A1"] = "PIZARRA vs FUTUROS MATBA"; estilo_titulo(ws["A1"], C_AZUL_OSC); ws.row_dimensions[1].height = 30
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols_f)
        ws["A2"] = f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
        ws["A2"].font = Font(italic=True, size=9, color="757575", name="Calibri"); ws["A2"].alignment = _center()
        write_headers(ws, 3, ["Posición", "Cereal", "Contrato", "Precio USD", "Variación", "Pizarra USD Hoy", "Diferencia %", "Señal", "Pizarra ARS"], C_AZUL_MED)
        row = 4
        for m in matba_rows:
            cer_matba = m["cereal"]; cer_simple = _mapear_cereal(cer_matba); merc = mercado.get(cer_simple) if cer_simple else None
            variacion = None
            try: variacion = m["variacion"]
            except: pass
            hoy_usd = None; dif = None; signal = "Sin dato"; piz_ars = None
            if merc and merc["pizarra"] and merc["dolar"]:
                piz_ars = merc["pizarra"]; hoy_usd = round(piz_ars / merc["dolar"], 2)
                if hoy_usd > 0 and m["precio"]:
                    dif = round(((m["precio"] - hoy_usd) / hoy_usd) * 100, 2)
                    signal = "Esperar" if dif > 5 else ("Vender Hoy" if dif < -2 else "Neutral")
            contrato = None
            try: contrato = m["mes"]
            except: pass
            bg = None
            if dif is not None: bg = C_VERDE_CLR if dif > 5 else (C_ROJO_CLR if dif < 0 else None)
            if bg is None and row % 2 == 0: bg = C_GRIS
            write_row(ws, row, [m["posicion"], cer_matba, contrato, m["precio"], variacion, hoy_usd, dif, signal, piz_ars], bg=bg, num_fmts=[None, None, None, "#,##0.00", "0.00", "#,##0.00", "0.00", None, "#,##0.00"]); row += 1
        set_col_widths(ws, [14, 22, 14, 14, 12, 16, 14, 14, 16]); ws.auto_filter.ref = f"A3:I{row}"; ws.freeze_panes = "A4"

    nivel = "admin" if es_admin else ("comercial" if ve_comercial else ("calidad" if ve_calidad else "form"))
    registrar_auditoria(conn, current_user.id, empresa_id, "exportacion_excel", f"Exportación nivel {nivel}", None)
    conn.close()

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        as_attachment=True,
        download_name=f"silos_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


@panel_bp.route("/seleccionar_empresa/<int:id>")
@login_required
def seleccionar_empresa(id):
    if not current_user.es_superadmin:
        return "No autorizado", 403
    session["empresa_contexto"] = id
    return redirect(url_for("panel.panel"))


@panel_bp.route("/cambiar_empresa")
@login_required
def cambiar_empresa():
    if not current_user.es_superadmin:
        return "No autorizado", 403
    session.pop("empresa_contexto", None)
    return redirect(url_for("panel.panel"))
